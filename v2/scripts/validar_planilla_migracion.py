"""
Revisa la planilla que devolvio bedelia, antes de tocar la base.

La planilla se va a completar a mano, y a mano siempre salen cedulas con puntos,
codigos de materia que no existen, alumnos con la materia 5 aprobada y la 1 sin
registrar. Este script encuentra todo eso y lo informa con hoja, fila y columna,
para poder devolverselo a administracion y repetir el ciclo las veces que haga
falta sin haber escrito nada en la base.

    python -m v2.scripts.validar_planilla_migracion carga_inicial.xlsx
    python -m v2.scripts.validar_planilla_migracion carga_inicial.xlsx --reporte errores.txt

ERROR   impide importar la fila.
AVISO   se puede importar, pero probablemente este mal y conviene mirarlo.

Sale con codigo 1 si hay errores. Los avisos no lo hacen fallar.

No lee ni escribe la base: trabaja solo sobre el archivo.
"""
import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook

from v2.scripts.generar_planilla_migracion import (
    ESTADOS_CARRERA, ESTADOS_HISTORIAL, ROLES_DOCENTE, SI_NO, TIPOS_PREVIATURA,
)
from v2.scripts.malla_inicial import normalizar

HOJAS = {
    "alumnos": "1-Alumnos",
    "docentes": "2-Docentes",
    "plan": "3-Plan de estudios",
    "previaturas": "4-Previaturas",
    "historial": "5-Historial",
    "dictado": "6-Dictado actual",
}

# Todas las hojas de datos tienen la nota en la fila 1 y el encabezado en la 2
PRIMERA_FILA = 3

ESTADOS_QUE_CUENTAN_COMO_TENIDA = {"APROBADA", "EXONERADA"}

RE_DOCUMENTO = re.compile(r"^\d{6,10}$")
RE_EMAIL = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


@dataclass
class Problema:
    nivel: str      # ERROR | AVISO
    hoja: str
    fila: Optional[int]
    mensaje: str

    def __str__(self) -> str:
        donde = f"{self.hoja}"
        if self.fila:
            donde += f" fila {self.fila}"
        return f"[{self.nivel}] {donde}: {self.mensaje}"


class Validador:
    def __init__(self, ruta: str):
        self.wb = load_workbook(ruta, data_only=True)
        self.problemas: List[Problema] = []

        # Catalogos que se van armando a medida que se leen las hojas
        # La clave de una materia es su codigo si lo tiene, y si no
        # "PROGRAMA::nombre". `Materia.codigo` es nullable en el modelo y la
        # malla que ya nos pasaron viene sin codigos, asi que exigirlos seria
        # inventar un requisito que el sistema no tiene.
        self.materias: Dict[str, dict] = {}          # clave -> datos
        self.por_codigo: Dict[str, str] = {}         # codigo -> clave
        self.por_nombre: Dict[Tuple[str, str], str] = {}   # (prog, nombre) -> clave
        self.nombres_ambiguos: Dict[str, List[str]] = defaultdict(list)
        self.programas: Set[str] = set()             # los que aparecen en el plan
        self.programas_de_alumno: Dict[str, Set[str]] = defaultdict(set)
        self.documentos_alumnos: Set[str] = set()
        self.documentos_docentes: Set[str] = set()
        self.previas_de: Dict[str, List[Tuple[str, str]]] = defaultdict(list)

    # ── Utilidades ───────────────────────────────────────────────────────────

    def error(self, hoja: str, fila: Optional[int], mensaje: str):
        self.problemas.append(Problema("ERROR", hoja, fila, mensaje))

    def aviso(self, hoja: str, fila: Optional[int], mensaje: str):
        self.problemas.append(Problema("AVISO", hoja, fila, mensaje))

    @staticmethod
    def _texto(valor) -> str:
        if valor is None:
            return ""
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    def _documento(self, valor) -> str:
        """Normaliza la cedula: solo digitos. Es la clave que une las hojas."""
        crudo = self._texto(valor)
        return re.sub(r"[.\-\s]", "", crudo)

    def _filas(self, clave: str):
        """Itera las filas con datos de una hoja, salteando las vacias."""
        nombre = HOJAS[clave]
        if nombre not in self.wb.sheetnames:
            self.error(nombre, None, "Falta la hoja. ¿Se borro o se renombro?")
            return
        ws = self.wb[nombre]
        for numero, fila in enumerate(
            ws.iter_rows(min_row=PRIMERA_FILA, values_only=True), start=PRIMERA_FILA
        ):
            if all(celda is None or str(celda).strip() == "" for celda in fila):
                continue
            yield numero, fila

    def _obligatorio(self, hoja: str, fila: int, valor, campo: str) -> bool:
        if not self._texto(valor):
            self.error(hoja, fila, f"Falta {campo}.")
            return False
        return True

    def _entero(self, hoja: str, fila: int, valor, campo: str,
                minimo: int, maximo: int) -> Optional[int]:
        texto = self._texto(valor)
        if not texto:
            return None
        try:
            numero = int(float(texto))
        except ValueError:
            self.error(hoja, fila, f"{campo} tiene que ser un numero: '{texto}'.")
            return None
        if not (minimo <= numero <= maximo):
            self.error(hoja, fila, f"{campo} fuera de rango ({minimo}-{maximo}): {numero}.")
            return None
        return numero

    def _resolver_materia(
        self, valor, programa: Optional[str] = None
    ) -> Tuple[Optional[str], Optional[str]]:
        """
        Encuentra la materia venga escrito el codigo o el nombre.

        La hoja de previaturas viene precargada con nombres, porque la malla que
        ya nos pasaron no trae codigos. Aceptar las dos formas evita pedirle a
        bedelia que traduzca 44 filas a mano.

        Devuelve (clave, motivo_del_fallo). Si el nombre existe en dos carreras
        y no se dice cual, es ambiguo y hay que aclararlo.
        """
        texto = self._texto(valor)
        if not texto:
            return None, "esta vacio"

        if texto.upper() in self.por_codigo:
            return self.por_codigo[texto.upper()], None

        normalizado = normalizar(texto)

        if programa:
            clave = self.por_nombre.get((normalizar(programa), normalizado))
            if clave:
                return clave, None

        candidatos = self.nombres_ambiguos.get(normalizado, [])
        if len(candidatos) == 1:
            return candidatos[0], None
        if len(candidatos) > 1:
            carreras = sorted(
                self.materias[c]["programa"] for c in candidatos
            )
            return None, (
                f"existe en mas de una carrera ({', '.join(carreras)}); "
                f"hay que aclarar cual en la columna Programa"
            )

        return None, f"no esta en la hoja '{HOJAS['plan']}', ni por codigo ni por nombre"

    def _opcion(self, hoja: str, fila: int, valor, campo: str,
                opciones: List[str]) -> Optional[str]:
        texto = self._texto(valor).upper()
        if not texto:
            self.error(hoja, fila, f"Falta {campo}.")
            return None
        if texto not in opciones:
            self.error(
                hoja, fila,
                f"{campo} dice '{texto}' y tiene que ser uno de: {', '.join(opciones)}.",
            )
            return None
        return texto

    # ── Hojas ────────────────────────────────────────────────────────────────

    def validar_plan(self):
        hoja = HOJAS["plan"]
        codigos_vistos: Dict[str, int] = {}
        nombres_vistos: Dict[Tuple[str, str], int] = {}
        incompletas: List[Tuple[int, str, str]] = []

        for fila, datos in self._filas("plan"):
            programa, codigo, nombre, semestre, creditos, dictando, _ = (
                list(datos) + [None] * 7
            )[:7]

            programa = self._texto(programa)
            codigo = self._texto(codigo).upper()
            nombre = self._texto(nombre)

            if not programa:
                self.error(hoja, fila, "Falta el programa.")
                continue
            if not nombre:
                self.error(hoja, fila, "Falta el nombre de la materia.")
                continue

            nombre_normalizado = normalizar(nombre)
            clave_nombre = (normalizar(programa), nombre_normalizado)
            if clave_nombre in nombres_vistos:
                self.error(
                    hoja, fila,
                    f"'{nombre}' ya aparece en '{programa}' en la fila "
                    f"{nombres_vistos[clave_nombre]}. Dentro de una carrera el "
                    f"nombre no se puede repetir.",
                )
                continue
            nombres_vistos[clave_nombre] = fila

            if codigo:
                if codigo in codigos_vistos:
                    self.error(
                        hoja, fila,
                        f"El codigo '{codigo}' ya aparece en la fila "
                        f"{codigos_vistos[codigo]}. Tiene que ser unico.",
                    )
                    continue
                codigos_vistos[codigo] = fila

            # El semestre y los creditos son obligatorios en el modelo; el
            # codigo no. Las filas precargadas desde la malla llegan sin
            # semestre ni creditos y son decenas: se acumulan y se informan
            # juntas, porque ochenta lineas identicas no le sirven a nadie.
            faltan = []
            if not self._texto(semestre):
                faltan.append("semestre")
            else:
                self._entero(hoja, fila, semestre, "el semestre del plan", 1, 20)
            if not self._texto(creditos):
                faltan.append("creditos")
            else:
                self._entero(hoja, fila, creditos, "los creditos", 0, 500)
            if faltan:
                incompletas.append((fila, nombre, " y ".join(faltan)))

            self._opcion(hoja, fila, dictando, "si se sigue dictando", SI_NO)

            clave = codigo or f"{programa}::{nombre_normalizado}"
            self.materias[clave] = {
                "programa": programa,
                "nombre": nombre,
                "codigo": codigo,
                "semestre": semestre,
                "fila": fila,
            }
            self.programas.add(programa)
            if codigo:
                self.por_codigo[codigo] = clave
            self.por_nombre[clave_nombre] = clave
            self.nombres_ambiguos[nombre_normalizado].append(clave)

        if incompletas:
            detalle = "; ".join(
                f"fila {fila}: '{nombre}' (falta {que})"
                for fila, nombre, que in incompletas[:15]
            )
            if len(incompletas) > 15:
                detalle += f"; y {len(incompletas) - 15} mas"
            self.error(
                hoja, None,
                f"{len(incompletas)} materias sin completar. Son las filas verdes, "
                f"que vienen de la malla ya definida y solo necesitan el semestre "
                f"del plan y los creditos: {detalle}",
            )

        if not self.materias:
            self.error(hoja, None, "No hay ninguna materia cargada.")

    def _mostrar(self, clave: str) -> str:
        """Como nombrar una materia en un mensaje: codigo si tiene, si no nombre."""
        datos = self.materias.get(clave, {})
        return datos.get("codigo") or datos.get("nombre") or clave

    def validar_previaturas(self):
        hoja = HOJAS["previaturas"]
        vistas: Set[Tuple[str, str]] = set()

        for fila, datos in self._filas("previaturas"):
            programa, entrada, entrada_previa, tipo, _ = (list(datos) + [None] * 5)[:5]

            if not self._texto(entrada) or not self._texto(entrada_previa):
                self.error(hoja, fila, "Falta la materia o la materia previa.")
                continue

            texto_programa = self._texto(programa)
            codigo, falla = self._resolver_materia(entrada, texto_programa)
            if codigo is None:
                self.error(hoja, fila, f"'{self._texto(entrada)}' {falla}.")
            codigo_previa, falla_previa = self._resolver_materia(
                entrada_previa, texto_programa
            )
            if codigo_previa is None:
                self.error(hoja, fila, f"'{self._texto(entrada_previa)}' {falla_previa}.")
            if codigo is None or codigo_previa is None:
                continue

            if codigo == codigo_previa:
                self.error(
                    hoja, fila,
                    f"'{self._mostrar(codigo)}' no puede ser previatura de si misma.",
                )
                continue

            if (self.materias[codigo]["programa"]
                    != self.materias[codigo_previa]["programa"]):
                self.error(
                    hoja, fila,
                    f"'{self._mostrar(codigo)}' y '{self._mostrar(codigo_previa)}' son "
                    f"de programas distintos. Una previatura tiene que ser dentro de "
                    f"la misma carrera.",
                )
                continue

            if (codigo, codigo_previa) in vistas:
                self.aviso(
                    hoja, fila,
                    f"'{self._mostrar(codigo)} requiere "
                    f"{self._mostrar(codigo_previa)}' esta repetido.",
                )
                continue
            vistas.add((codigo, codigo_previa))

            self._opcion(hoja, fila, tipo, "el tipo de previatura", TIPOS_PREVIATURA)
            self.previas_de[codigo].append((codigo_previa, str(fila)))

        self._buscar_ciclos()

    def _buscar_ciclos(self):
        """
        Un ciclo hace que nadie pueda cursar ninguna de esas materias nunca.
        Mejor encontrarlo en la planilla que despues de importar.
        """
        hoja = HOJAS["previaturas"]
        visitados: Set[str] = set()

        for inicio in list(self.previas_de.keys()):
            if inicio in visitados:
                continue
            pila = [(inicio, [inicio])]
            while pila:
                actual, camino = pila.pop()
                for siguiente, _ in self.previas_de.get(actual, []):
                    if siguiente in camino:
                        ciclo = camino[camino.index(siguiente):] + [siguiente]
                        self.error(
                            hoja, None,
                            "Ciclo de previaturas: "
                            + " requiere ".join(self._mostrar(c) for c in ciclo)
                            + ". Hay que sacar uno de esos requisitos.",
                        )
                        pila = []
                        break
                    pila.append((siguiente, camino + [siguiente]))
                visitados.add(actual)

    def validar_alumnos(self):
        hoja = HOJAS["alumnos"]
        documentos_por_fila: Dict[Tuple[str, str], int] = {}
        nombres_por_documento: Dict[str, str] = {}
        anio_actual = datetime.now().year

        for fila, datos in self._filas("alumnos"):
            (documento, apellido, nombre, email, _personal, _tel,
             _nacimiento, programa, anio, estado, _obs) = (list(datos) + [None] * 11)[:11]

            documento = self._documento(documento)
            if not documento:
                self.error(hoja, fila, "Falta el documento.")
                continue
            if not RE_DOCUMENTO.match(documento):
                self.error(
                    hoja, fila,
                    f"El documento '{documento}' no parece una cedula. "
                    f"Solo numeros, sin puntos ni guiones.",
                )
                continue

            self._obligatorio(hoja, fila, apellido, "el apellido")
            self._obligatorio(hoja, fila, nombre, "el nombre")

            texto_email = self._texto(email)
            if texto_email and not RE_EMAIL.match(texto_email):
                self.error(hoja, fila, f"El email '{texto_email}' esta mal escrito.")

            # Antes de cualquier `continue`: una misma persona escrita distinto en
            # dos filas es casi siempre un typo, y se pierde si se saltea la fila
            nombre_completo = f"{self._texto(apellido)}, {self._texto(nombre)}"
            anterior = nombres_por_documento.get(documento)
            if anterior and anterior.lower() != nombre_completo.lower():
                self.aviso(
                    hoja, fila,
                    f"El documento {documento} figura como '{anterior}' y tambien "
                    f"como '{nombre_completo}'. ¿Es la misma persona?",
                )
            nombres_por_documento[documento] = nombre_completo

            programa = self._texto(programa)
            if not programa:
                self.error(hoja, fila, "Falta el programa.")
                continue

            # Contra el plan y no contra la base: asi bedelia puede correr el
            # validador sin acceso a nada. El importador despues chequea que el
            # programa exista de verdad.
            if programa not in self.programas:
                self.error(
                    hoja, fila,
                    f"El programa '{programa}' no aparece en la hoja "
                    f"'{HOJAS['plan']}'. Si es una carrera nueva, hay que cargar "
                    f"sus materias ahi primero.",
                )
                continue

            clave = (documento, programa)
            if clave in documentos_por_fila:
                self.error(
                    hoja, fila,
                    f"El documento {documento} ya figura en '{programa}' "
                    f"en la fila {documentos_por_fila[clave]}.",
                )
                continue
            documentos_por_fila[clave] = fila

            self._entero(hoja, fila, anio, "el año de ingreso", 1990, anio_actual + 1)
            self._opcion(hoja, fila, estado, "el estado en la carrera", ESTADOS_CARRERA)

            self.documentos_alumnos.add(documento)
            self.programas_de_alumno[documento].add(programa)

        if not self.documentos_alumnos:
            self.error(hoja, None, "No hay ningun alumno cargado.")

    def validar_docentes(self):
        hoja = HOJAS["docentes"]
        vistos: Dict[str, int] = {}

        for fila, datos in self._filas("docentes"):
            documento, apellido, nombre, email, _personal, _tel, activo, _obs = (
                list(datos) + [None] * 8
            )[:8]

            documento = self._documento(documento)
            if not documento:
                self.error(hoja, fila, "Falta el documento.")
                continue
            if not RE_DOCUMENTO.match(documento):
                self.error(hoja, fila, f"El documento '{documento}' no parece una cedula.")
                continue
            if documento in vistos:
                self.error(
                    hoja, fila,
                    f"El documento {documento} ya aparece en la fila {vistos[documento]}.",
                )
                continue
            vistos[documento] = fila

            self._obligatorio(hoja, fila, apellido, "el apellido")
            self._obligatorio(hoja, fila, nombre, "el nombre")

            texto_email = self._texto(email)
            if texto_email and not RE_EMAIL.match(texto_email):
                self.error(hoja, fila, f"El email '{texto_email}' esta mal escrito.")

            self._opcion(hoja, fila, activo, "si sigue activo", SI_NO)
            self.documentos_docentes.add(documento)

    def validar_historial(self):
        hoja = HOJAS["historial"]
        vistas: Dict[Tuple[str, str], int] = {}
        anio_actual = datetime.now().year
        # documento -> {codigo: estado}, para el chequeo de cadenas
        estados: Dict[str, Dict[str, str]] = defaultdict(dict)

        for fila, datos in self._filas("historial"):
            documento, programa, codigo, estado, nota, anio, semestre, _obs = (
                list(datos) + [None] * 8
            )[:8]

            documento = self._documento(documento)

            if not documento or not self._texto(codigo):
                self.error(hoja, fila, "Faltan el documento o la materia.")
                continue

            if documento not in self.documentos_alumnos:
                self.error(
                    hoja, fila,
                    f"El documento {documento} no esta en la hoja '{HOJAS['alumnos']}'. "
                    f"Todo alumno del historial tiene que estar cargado ahi primero.",
                )
                continue

            resuelto, falla = self._resolver_materia(codigo, self._texto(programa))
            if resuelto is None:
                self.error(hoja, fila, f"'{self._texto(codigo)}' {falla}.")
                continue
            codigo = resuelto

            clave = (documento, codigo)
            if clave in vistas:
                self.error(
                    hoja, fila,
                    f"{documento} ya tiene una fila para "
                    f"'{self._mostrar(codigo)}' (fila {vistas[clave]}). "
                    f"Una sola por alumno y materia.",
                )
                continue
            vistas[clave] = fila

            # La materia tiene que ser de una carrera en la que el alumno este
            programa_materia = self.materias[codigo]["programa"]
            if programa_materia not in self.programas_de_alumno[documento]:
                self.error(
                    hoja, fila,
                    f"'{self._mostrar(codigo)}' es de '{programa_materia}', pero "
                    f"{documento} no figura inscripto en esa carrera.",
                )
                continue

            texto_programa = self._texto(programa)
            if texto_programa and texto_programa != programa_materia:
                self.aviso(
                    hoja, fila,
                    f"Dice '{texto_programa}' pero '{self._mostrar(codigo)}' es de "
                    f"'{programa_materia}'. Se toma el de la materia.",
                )

            valor_estado = self._opcion(hoja, fila, estado, "el estado", ESTADOS_HISTORIAL)
            if valor_estado is None:
                continue
            estados[documento][codigo] = valor_estado

            if self._texto(nota):
                texto_nota = self._texto(nota).replace(",", ".")
                try:
                    numero = float(texto_nota)
                    if not (0 <= numero <= 100):
                        self.error(hoja, fila, f"La nota {numero} esta fuera de 0 a 100.")
                except ValueError:
                    self.error(hoja, fila, f"La nota '{texto_nota}' no es un numero.")

            if self._texto(anio):
                self._entero(hoja, fila, anio, "el año", 1990, anio_actual + 1)
            elif valor_estado in ESTADOS_QUE_CUENTAN_COMO_TENIDA:
                self.aviso(
                    hoja, fila,
                    "Materia aprobada sin año. Se puede importar, pero la "
                    "escolaridad va a salir sin fecha.",
                )

            if self._texto(semestre):
                self._entero(hoja, fila, semestre, "el semestre", 1, 2)

        self._revisar_cadenas(estados)

    def _revisar_cadenas(self, estados: Dict[str, Dict[str, str]]):
        """
        El chequeo que mas dolores de cabeza ahorra.

        Si un alumno tiene una materia aprobada pero su previatura no figura,
        el portal lo va a bloquear el dia uno para inscribirse a lo que sigue, y
        bedelia va a reportar que el sistema esta roto. Casi siempre es que el
        historial viejo quedo incompleto, no un error real del alumno.

        Va como AVISO y no como ERROR porque tambien puede ser legitimo: una
        equivalencia, un cambio de plan, o una excepcion que despues se carga a
        mano. Pero hay que mirarlo caso por caso antes de importar.
        """
        hoja = HOJAS["historial"]
        for documento, materias_alumno in sorted(estados.items()):
            for codigo, estado in sorted(materias_alumno.items()):
                if estado not in ESTADOS_QUE_CUENTAN_COMO_TENIDA:
                    continue
                for previa, _ in self.previas_de.get(codigo, []):
                    estado_previa = materias_alumno.get(previa)
                    if estado_previa in ESTADOS_QUE_CUENTAN_COMO_TENIDA:
                        continue
                    detalle = (
                        f"figura como {estado_previa}" if estado_previa
                        else "no figura en el historial"
                    )
                    self.aviso(
                        hoja, None,
                        f"{documento}: tiene '{self._mostrar(codigo)}' {estado} pero "
                        f"su previatura '{self._mostrar(previa)}' {detalle}. Revisar: "
                        f"si es historial viejo que falta, agregarlo; si el alumno la "
                        f"debe de verdad, va a necesitar una excepcion de previatura.",
                    )

    def validar_dictado(self):
        hoja = HOJAS["dictado"]
        anio_actual = datetime.now().year
        vistos: Set[Tuple[str, int, int, str]] = set()

        for fila, datos in self._filas("dictado"):
            (programa, codigo, anio, semestre, documento, rol,
             _horario, _salon, cupo, _obs) = (list(datos) + [None] * 10)[:10]

            documento = self._documento(documento)

            if not self._texto(codigo):
                self.error(hoja, fila, "Falta la materia.")
                continue
            resuelto, falla = self._resolver_materia(codigo, self._texto(programa))
            if resuelto is None:
                self.error(hoja, fila, f"'{self._texto(codigo)}' {falla}.")
                continue
            codigo = resuelto

            if documento and documento not in self.documentos_docentes:
                self.error(
                    hoja, fila,
                    f"El documento {documento} no esta en la hoja "
                    f"'{HOJAS['docentes']}'.",
                )
            elif not documento:
                self.error(hoja, fila, "Falta el documento del docente.")

            valor_anio = self._entero(hoja, fila, anio, "el año", 1990, anio_actual + 2)
            valor_semestre = self._entero(hoja, fila, semestre, "el semestre", 1, 2)
            self._opcion(hoja, fila, rol, "el rol del docente", ROLES_DOCENTE)

            if self._texto(cupo):
                self._entero(hoja, fila, cupo, "el cupo maximo", 1, 500)

            if valor_anio and valor_semestre and documento:
                clave = (codigo, valor_anio, valor_semestre, documento)
                if clave in vistos:
                    self.aviso(
                        hoja, fila,
                        f"{documento} ya figura en '{self._mostrar(codigo)}' "
                        f"{valor_anio}/S{valor_semestre}.",
                    )
                vistos.add(clave)

    # ── Ejecucion ────────────────────────────────────────────────────────────

    def correr(self) -> List[Problema]:
        # El orden importa: cada hoja arma los catalogos que usa la siguiente
        self.validar_plan()
        self.validar_previaturas()
        self.validar_alumnos()
        self.validar_docentes()
        self.validar_historial()
        self.validar_dictado()
        return self.problemas


def _resumen(validador: Validador) -> List[str]:
    return [
        "Resumen de lo leido:",
        f"  Materias en el plan      {len(validador.materias)}",
        f"  Previaturas              {sum(len(v) for v in validador.previas_de.values())}",
        f"  Alumnos                  {len(validador.documentos_alumnos)}",
        f"  Docentes                 {len(validador.documentos_docentes)}",
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archivo", help="Planilla .xlsx devuelta por bedelia")
    parser.add_argument("--reporte", help="Guarda el detalle en un .txt para reenviar")
    args = parser.parse_args()

    validador = Validador(args.archivo)
    problemas = validador.correr()

    errores = [p for p in problemas if p.nivel == "ERROR"]
    avisos = [p for p in problemas if p.nivel == "AVISO"]

    lineas = _resumen(validador) + [""]

    if errores:
        lineas.append(f"ERRORES ({len(errores)}) - hay que corregirlos para importar:")
        lineas += [f"  {p}" for p in errores]
        lineas.append("")
    if avisos:
        lineas.append(f"AVISOS ({len(avisos)}) - revisar, no bloquean:")
        lineas += [f"  {p}" for p in avisos]
        lineas.append("")

    if not errores and not avisos:
        lineas.append("Todo bien. La planilla esta lista para importar.")
    elif not errores:
        lineas.append("Sin errores. Mirá los avisos y despues se puede importar.")

    texto = "\n".join(lineas)
    print(texto)

    if args.reporte:
        with open(args.reporte, "w", encoding="utf-8") as archivo:
            archivo.write(texto + "\n")
        print(f"\nReporte guardado en {args.reporte}")

    return 1 if errores else 0


if __name__ == "__main__":
    sys.exit(main())
