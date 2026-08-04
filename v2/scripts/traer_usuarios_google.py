"""
Precarga en la planilla las personas que ya existen en Google Workspace.

Le ahorra a bedelia escribir a mano cada nombre y cada correo, y sobre todo
evita que los escriba distinto: el correo institucional es con lo que la persona
entra al portal, asi que un error de tipeo ahi la deja afuera.

    python -m v2.scripts.traer_usuarios_google carga_inicial.xlsx
    python -m v2.scripts.traer_usuarios_google carga_inicial.xlsx --solo-ver

Va contra el webhook de n8n que ya se usa para administrar las cuentas, asi que
no hace falta ninguna credencial nueva.

LO QUE NO TRAE
--------------
La CEDULA. Google no la tiene cargada en estas cuentas, y es la clave que une
todas las hojas: esa columna hay que completarla igual.

EL ROL, si el workflow de n8n viene simplificado. Google clasifica a la gente
por unidad organizativa (/Alumnos, /Equipo Docente, /Administración y Ventas),
pero el webhook `getManyUsersGoogle` tiene activada la opcion "Simplify" y no
devuelve `orgUnitPath`. Sin ese dato no se puede separar alumnos de docentes, y
el script deja todo en la hoja de alumnos marcado para revisar.

Para que salga clasificado, en n8n: abrir el workflow `getManyUsersGoogle` y
desactivar "Simplify" en el nodo de Google Workspace. Con eso viene la OU y este
script reparte solo. (El otro camino, el workflow `google-user-ou`, hoy esta
inactivo y devuelve 404.)
"""
import argparse
import sys
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from external_services.google.google_service import GoogleWorkspaceService
from v2.auth.n8n_ou_client import N8nOUClient
from v2.models.enums import RolUsuario
from v2.scripts.validar_planilla_migracion import HOJAS, PRIMERA_FILA

RELLENO_GOOGLE = PatternFill("solid", fgColor="DDEBF7")
AVISO_SIN_OU = "VERIFICAR ROL - Google no devolvio la unidad organizativa"


def _usuarios_de(respuesta) -> List[dict]:
    """La lista de usuarios, venga como venga envuelta desde n8n."""
    if isinstance(respuesta, list):
        crudos = respuesta
    elif isinstance(respuesta, dict):
        crudos = next(
            (respuesta[clave] for clave in ("users", "data", "items", "result")
             if isinstance(respuesta.get(clave), list)),
            [],
        )
    else:
        crudos = []

    # n8n a veces envuelve cada item en {"json": {...}}
    return [
        item.get("json", item) if isinstance(item, dict) else item
        for item in crudos
    ]


def traer_directorio() -> List[dict]:
    """Trae todas las cuentas, paginando si hace falta."""
    servicio = GoogleWorkspaceService()
    usuarios: List[dict] = []
    token: Optional[str] = None

    while True:
        respuesta = servicio.list_google_accounts(max_results=500, page_token=token)
        pagina = _usuarios_de(respuesta)
        usuarios.extend(pagina)

        token = respuesta.get("nextPageToken") if isinstance(respuesta, dict) else None
        if not token or not pagina:
            break

    return usuarios


def clasificar(usuarios: List[dict]) -> Dict[Optional[RolUsuario], List[dict]]:
    """Reparte por rol segun la unidad organizativa. None = no se sabe."""
    por_rol: Dict[Optional[RolUsuario], List[dict]] = {}
    for usuario in usuarios:
        rol = N8nOUClient.ou_to_rol(usuario.get("orgUnitPath"))
        por_rol.setdefault(rol, []).append(usuario)
    return por_rol


def _emails_ya_cargados(ws, columna: int) -> set:
    return {
        str(ws.cell(row=fila, column=columna).value).strip().lower()
        for fila in range(PRIMERA_FILA, ws.max_row + 1)
        if ws.cell(row=fila, column=columna).value
    }


def _primera_fila_libre(ws) -> int:
    fila = PRIMERA_FILA
    while any(
        ws.cell(row=fila, column=columna).value not in (None, "")
        for columna in range(1, 12)
    ):
        fila += 1
    return fila


def _datos_persona(usuario: dict) -> tuple:
    nombre = usuario.get("name") or {}
    return (
        (nombre.get("familyName") or "").strip(),
        (nombre.get("givenName") or "").strip(),
        (usuario.get("primaryEmail") or "").strip(),
        bool(usuario.get("suspended")),
    )


def escribir_alumnos(ws, usuarios: List[dict], sin_ou: bool) -> int:
    """Hoja 1-Alumnos. El documento y el programa los pone bedelia."""
    ya = _emails_ya_cargados(ws, 4)
    fila = _primera_fila_libre(ws)
    escritos = 0

    for usuario in usuarios:
        apellido, nombre, email, suspendida = _datos_persona(usuario)
        if not email or email.lower() in ya:
            continue

        observaciones = []
        if sin_ou:
            observaciones.append(AVISO_SIN_OU)
        if suspendida:
            observaciones.append("cuenta de Google suspendida")

        valores = [
            "", apellido, nombre, email, "", "", "",
            "", "", "", "; ".join(observaciones),
        ]
        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=columna, value=valor or None)
            if columna in (2, 3, 4):
                celda.fill = RELLENO_GOOGLE

        fila += 1
        escritos += 1

    return escritos


def escribir_docentes(ws, usuarios: List[dict]) -> int:
    """Hoja 2-Docentes."""
    ya = _emails_ya_cargados(ws, 4)
    fila = _primera_fila_libre(ws)
    escritos = 0

    for usuario in usuarios:
        apellido, nombre, email, suspendida = _datos_persona(usuario)
        if not email or email.lower() in ya:
            continue

        valores = [
            "", apellido, nombre, email, "", "",
            "NO" if suspendida else "SI",
            "cuenta de Google suspendida" if suspendida else "",
        ]
        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=columna, value=valor or None)
            if columna in (2, 3, 4):
                celda.fill = RELLENO_GOOGLE

        fila += 1
        escritos += 1

    return escritos


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("archivo", help="Planilla .xlsx a completar")
    parser.add_argument(
        "--solo-ver", action="store_true",
        help="Muestra lo que traeria sin escribir el archivo",
    )
    args = parser.parse_args()

    print("Consultando el directorio de Google via n8n...")
    try:
        usuarios = traer_directorio()
    except Exception as e:
        print(f"No se pudo consultar: {type(e).__name__}: {e}")
        return 1

    print(f"Cuentas encontradas: {len(usuarios)}")
    if not usuarios:
        return 1

    por_rol = clasificar(usuarios)
    sin_ou = list(por_rol.get(None, []))
    hay_clasificacion = any(rol is not None for rol in por_rol)

    print("\nPor unidad organizativa:")
    for rol, gente in sorted(
        por_rol.items(), key=lambda par: (par[0].value if par[0] else "")
    ):
        etiqueta = rol.value if rol else "sin OU / no reconocida"
        print(f"  {etiqueta:<28} {len(gente)}")

    suspendidas = sum(1 for u in usuarios if u.get("suspended"))
    if suspendidas:
        print(f"\n  {suspendidas} cuentas suspendidas (se marcan en Observaciones)")

    if not hay_clasificacion:
        print(
            "\nATENCION: Google no devolvio la unidad organizativa, asi que no se "
            "puede separar alumnos de docentes.\n"
            "Todas las personas van a '1-Alumnos' marcadas para revisar.\n"
            "\nPara que salga clasificado: en n8n, abrir el workflow "
            "'getManyUsersGoogle' y desactivar\n'Simplify' en el nodo de Google "
            "Workspace. Despues volver a correr este script."
        )

    if args.solo_ver:
        print("\n--solo-ver: no se escribio nada.")
        return 0

    wb = load_workbook(args.archivo)
    for clave in ("alumnos", "docentes"):
        if HOJAS[clave] not in wb.sheetnames:
            print(f"La planilla no tiene la hoja '{HOJAS[clave]}'.")
            return 1

    if hay_clasificacion:
        alumnos = por_rol.get(RolUsuario.ESTUDIANTE, [])
        docentes = por_rol.get(RolUsuario.DOCENTE, [])
        # Los administrativos no tienen hoja propia: no participan de la
        # escolaridad. Se cargan aparte cuando se les da acceso al portal.
        administrativos = por_rol.get(RolUsuario.ADMINISTRATIVO, [])
        if administrativos:
            print(
                f"\n  {len(administrativos)} administrativos: no van en la planilla, "
                f"se les da acceso desde el portal."
            )
        # Los de OU no reconocida van a alumnos, marcados
        alumnos = alumnos + sin_ou
    else:
        alumnos, docentes = usuarios, []

    escritos_alumnos = escribir_alumnos(
        wb[HOJAS["alumnos"]], alumnos, sin_ou=not hay_clasificacion
    )
    escritos_docentes = escribir_docentes(wb[HOJAS["docentes"]], docentes)

    wb.save(args.archivo)

    print(f"\nEscrito en {args.archivo}:")
    print(f"  1-Alumnos    +{escritos_alumnos}")
    print(f"  2-Docentes   +{escritos_docentes}")
    print(
        "\nFalta completar a mano la CEDULA de cada persona (Google no la tiene) "
        "y el programa\nde cada alumno. Los nombres y correos ya estan."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
