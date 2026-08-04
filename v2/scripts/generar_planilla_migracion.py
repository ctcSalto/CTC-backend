"""
Genera la planilla que completa bedelia para la carga inicial del portal.

La idea es que administracion no escriba texto libre en ningun campo que despues
haya que interpretar. Todo lo que ya sabemos va precargado (programas, materias,
codigos) y todo lo que es un valor cerrado va con lista desplegable. Lo que
quede mal igual lo agarra `validar_planilla_migracion.py` antes de tocar la base.

La clave de todo es el DOCUMENTO (cedula). No los nombres, que se escriben de
mil formas, ni los ids internos, que bedelia no conoce.

    python -m v2.scripts.generar_planilla_migracion
    python -m v2.scripts.generar_planilla_migracion --salida C:\\ruta\\planilla.xlsx

Lee la base para precargar los catalogos. No escribe nada en ella.
"""
import argparse
import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from database.database import get_db_session
from v2.models.materia import Materia
from v2.models.previatura import Previatura
from v2.models.programa import Programa
from v2.scripts.malla_inicial import (
    materias_por_programa, normalizar, previaturas_por_programa,
)

# ── Vocabulario de bedelia ───────────────────────────────────────────────────
# Deliberadamente mas chico que EstadoInscripcionMateria: administracion no
# lleva registro de "perdido por inasistencia" de hace cuatro años, y ofrecer
# ocho opciones donde alcanzan cinco solo genera datos inconsistentes.
# La traduccion a los enums reales la hace el importador.
ESTADOS_HISTORIAL = [
    "APROBADA",    # aprobo, por examen o por nota de curso
    "EXONERADA",   # exonero, no rindio examen
    "A_EXAMEN",    # gano derecho a examen y todavia no lo rindio
    "CURSANDO",    # la esta cursando ahora
    "RECURSA",     # la curso y no la aprobo, tiene que volver a cursarla
]

ESTADOS_CARRERA = ["ACTIVA", "SUSPENDIDA", "COMPLETADA", "BAJA"]
TIPOS_PREVIATURA = ["APROBADA", "EXONERADA"]
ROLES_DOCENTE = ["TITULAR", "ADJUNTO", "ASISTENTE"]
SI_NO = ["SI", "NO"]

# ── Estilos ──────────────────────────────────────────────────────────────────

AZUL = "1F4E79"
GRIS = "F2F2F2"
AMARILLO = "FFF2CC"

FUENTE_TITULO = Font(bold=True, color="FFFFFF", size=11)
RELLENO_TITULO = PatternFill("solid", fgColor=AZUL)
RELLENO_PRECARGADO = PatternFill("solid", fgColor=GRIS)
RELLENO_COMPLETAR = PatternFill("solid", fgColor=AMARILLO)
# Lo que viene de la malla ya definida pero todavia no esta en la base: hay que
# distinguirlo de lo que el sistema ya tiene cargado.
RELLENO_MALLA = PatternFill("solid", fgColor="E2EFDA")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


class Columna:
    """Una columna de la planilla: como se titula, que ancho y que se acepta."""

    def __init__(
        self,
        titulo: str,
        ancho: int = 18,
        opciones: Optional[List[str]] = None,
        obligatoria: bool = False,
        ayuda: str = "",
    ):
        self.titulo = titulo
        self.ancho = ancho
        self.opciones = opciones
        self.obligatoria = obligatoria
        self.ayuda = ayuda


FILAS_VALIDACION = 3000  # hasta donde se extienden los desplegables


def _armar_hoja(
    wb: Workbook, nombre: str, columnas: List[Columna], nota: str = ""
) -> Worksheet:
    """Crea la hoja con su encabezado, anchos, desplegables y panel congelado."""
    ws = wb.create_sheet(nombre)
    fila_encabezado = 1

    if nota:
        ws.cell(row=1, column=1, value=nota)
        ws.cell(row=1, column=1).font = Font(italic=True, size=10, color="806000")
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=max(len(columnas), 1)
        )
        fila_encabezado = 2

    for indice, columna in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_encabezado, column=indice, value=columna.titulo)
        celda.font = FUENTE_TITULO
        celda.fill = RELLENO_TITULO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE
        ws.column_dimensions[get_column_letter(indice)].width = columna.ancho

        if columna.ayuda:
            celda.comment = None  # los comentarios de openpyxl molestan al abrir
            # La ayuda va en la hoja LEEME, no como comentario flotante

        if columna.opciones:
            validacion = DataValidation(
                type="list",
                formula1='"{}"'.format(",".join(columna.opciones)),
                allow_blank=not columna.obligatoria,
                showDropDown=False,
            )
            validacion.error = "Elegí un valor de la lista."
            validacion.errorTitle = "Valor no permitido"
            ws.add_data_validation(validacion)
            letra = get_column_letter(indice)
            validacion.add(f"{letra}{fila_encabezado + 1}:{letra}{FILAS_VALIDACION}")

    ws.freeze_panes = ws.cell(row=fila_encabezado + 1, column=1)
    ws.row_dimensions[fila_encabezado].height = 30
    return ws


def _escribir_fila(ws: Worksheet, fila: int, valores: list,
                   relleno: Optional[PatternFill] = None):
    for indice, valor in enumerate(valores, start=1):
        celda = ws.cell(row=fila, column=indice, value=valor)
        celda.border = BORDE
        if relleno is not None:
            celda.fill = relleno


# ── Hojas ────────────────────────────────────────────────────────────────────

def _hoja_leeme(wb: Workbook, programas: List[Programa]) -> None:
    ws = wb.create_sheet("LEEME")
    ws.column_dimensions["A"].width = 110

    lineas = [
        ("Carga inicial del Portal Academico", "titulo"),
        ("", ""),
        ("Esta planilla es la unica fuente de datos para arrancar el portal. "
         "Lo que no este aca, el sistema no lo va a saber.", ""),
        ("", ""),
        ("COMO COMPLETARLA", "seccion"),
        ("1. Las celdas GRISES ya vienen cargadas desde el sistema. "
         "Si algo esta mal, corregilo; si falta, agregalo abajo.", ""),
        ("2. Las columnas con flechita tienen lista de opciones. Usá la lista, "
         "no escribas a mano: si escribis 'aprobo' en vez de elegir 'APROBADA', "
         "esa fila no se va a poder importar.", ""),
        ("3. El DOCUMENTO (cedula) es lo que une todo. Escribilo siempre igual, "
         "solo numeros, sin puntos ni guiones. Si una persona aparece en dos "
         "hojas, tiene que ser el mismo numero en las dos.", ""),
        ("4. No borres ni renombres las hojas, ni cambies el orden de las "
         "columnas. Agregar filas al final esta perfecto.", ""),
        ("", ""),
        ("EN QUE ORDEN CONVIENE LLENARLA", "seccion"),
        ("Primero '3-Plan de estudios' y '4-Previaturas': definen las materias "
         "que despues se eligen en las otras hojas.", ""),
        ("Despues '1-Alumnos' y '2-Docentes': las personas.", ""),
        ("Al final '5-Historial' y '6-Dictado actual', que usan las dos cosas "
         "anteriores.", ""),
        ("", ""),
        ("LA HOJA IMPORTANTE ES '5-Historial'", "seccion"),
        ("Es la que dice, para cada alumno, como viene con cada materia. "
         "De ahi sale la escolaridad y de ahi sale a que se puede inscribir.", ""),
        ("", ""),
        ("Una fila por materia que el alumno TENGA ALGO. Si nunca la curso, "
         "no pongas la fila: se asume que la debe.", ""),
        ("", ""),
        ("  APROBADA   la aprobo (por examen o por nota de curso)", "mono"),
        ("  EXONERADA  exonero, no rindio examen", "mono"),
        ("  A_EXAMEN   gano el derecho a examen y todavia no lo rindio", "mono"),
        ("  CURSANDO   la esta cursando ahora mismo", "mono"),
        ("  RECURSA    la curso, no la aprobo, tiene que volver a cursarla", "mono"),
        ("", ""),
        ("OJO CON LAS MATERIAS APROBADAS HACE AÑOS", "seccion"),
        ("Si un alumno va por 5to semestre, necesitamos el historial COMPLETO, "
         "no solo lo de este año. El sistema controla previaturas: si figura "
         "Programacion 3 aprobada pero no figura Programacion 1, va a bloquear "
         "a ese alumno para inscribirse a lo que sigue.", ""),
        ("", ""),
        ("Si de una materia vieja no tenes la nota, dejá la nota vacia y poné "
         "igual el estado. El estado es lo que importa; la nota es informativa.", ""),
        ("", ""),
        ("DUDAS", "seccion"),
        ("Cualquier caso raro (alumno que cambio de plan, materia que ya no se "
         "dicta, equivalencias de otra institucion) anotalo en la columna "
         "'Observaciones' y lo vemos juntos. No lo fuerces a entrar.", ""),
        ("", ""),
        (f"Planilla generada el {datetime.now().strftime('%d/%m/%Y')}.", "pie"),
    ]

    fila = 1
    for texto, estilo in lineas:
        celda = ws.cell(row=fila, column=1, value=texto)
        celda.alignment = Alignment(wrap_text=True, vertical="top")
        if estilo == "titulo":
            celda.font = Font(bold=True, size=16, color=AZUL)
        elif estilo == "seccion":
            celda.font = Font(bold=True, size=12, color=AZUL)
        elif estilo == "mono":
            celda.font = Font(name="Consolas", size=10)
        elif estilo == "pie":
            celda.font = Font(italic=True, size=9, color="808080")
        else:
            celda.font = Font(size=11)
        fila += 1

    if programas:
        ws.cell(row=fila + 1, column=1, value="Programas en el sistema:").font = Font(
            bold=True, size=11
        )
        for indice, programa in enumerate(programas, start=fila + 2):
            ws.cell(row=indice, column=1, value=f"  - {programa.nombre}")


def _hoja_alumnos(wb: Workbook, nombres_programas: List[str]) -> None:
    columnas = [
        Columna("DOCUMENTO*", 16, obligatoria=True),
        Columna("Apellido*", 20, obligatoria=True),
        Columna("Nombre*", 20, obligatoria=True),
        Columna("Email institucional", 30),
        Columna("Email personal", 30),
        Columna("Telefono", 16),
        Columna("Fecha nacimiento\n(dd/mm/aaaa)", 16),
        Columna("Programa*", 32, opciones=nombres_programas, obligatoria=True),
        Columna("Año de ingreso*", 14, obligatoria=True),
        Columna("Estado en la carrera*", 18, opciones=ESTADOS_CARRERA, obligatoria=True),
        Columna("Observaciones", 40),
    ]
    _armar_hoja(
        wb, "1-Alumnos", columnas,
        nota="Una fila por alumno. Si un alumno cursa dos carreras, una fila por carrera "
             "(mismo documento repetido). El email institucional es con el que entra al "
             "portal: si no lo tiene todavia, dejalo vacio y se lo creamos.",
    )


def _hoja_docentes(wb: Workbook) -> None:
    columnas = [
        Columna("DOCUMENTO*", 16, obligatoria=True),
        Columna("Apellido*", 20, obligatoria=True),
        Columna("Nombre*", 20, obligatoria=True),
        Columna("Email institucional", 30),
        Columna("Email personal", 30),
        Columna("Telefono", 16),
        Columna("¿Sigue activo?*", 14, opciones=SI_NO, obligatoria=True),
        Columna("Observaciones", 40),
    ]
    _armar_hoja(
        wb, "2-Docentes", columnas,
        nota="Todos los docentes, incluidos los que ya no dan clase (poné 'NO' en activo): "
             "hacen falta para el historial de quien dicto que.",
    )


def _hoja_plan(wb: Workbook, session: Session, nombres_programas: List[str]) -> None:
    columnas = [
        Columna("Programa*", 32, opciones=nombres_programas, obligatoria=True),
        Columna("Codigo de materia\n(opcional)", 18),
        Columna("Nombre de la materia*", 36, obligatoria=True),
        Columna("Semestre del plan*", 14, obligatoria=True),
        Columna("Creditos*", 12, obligatoria=True),
        Columna("¿Se sigue dictando?*", 16, opciones=SI_NO, obligatoria=True),
        Columna("Observaciones", 40),
    ]
    ws = _armar_hoja(
        wb, "3-Plan de estudios", columnas,
        nota="GRIS = ya esta en el sistema. VERDE = sale de la malla que ya nos pasaron; "
             "falta completarle el SEMESTRE y los CREDITOS. Revisá todo y agregá abajo "
             "lo que falte. Lo que identifica a la materia es el NOMBRE: no puede "
             "repetirse dentro de una misma carrera. El codigo es opcional, pero si lo "
             "ponés tiene que ser unico.",
    )

    materias = session.exec(
        select(Materia, Programa)
        .join(Programa, Materia.programa_id == Programa.id)
        .order_by(Programa.nombre, Materia.semestre, Materia.nombre)
    ).all()

    fila = 3
    ya_en_base = set()
    for materia, programa in materias:
        _escribir_fila(ws, fila, [
            programa.nombre, materia.codigo, materia.nombre,
            materia.semestre, materia.creditos,
            "SI" if materia.activo else "NO", "",
        ], relleno=RELLENO_PRECARGADO)
        ya_en_base.add((programa.nombre, normalizar(materia.nombre)))
        fila += 1

    # Las materias que la malla menciona y todavia no existen. Van con el nombre
    # puesto y el resto vacio: bedelia completa codigo, semestre y creditos, que
    # es bastante menos trabajo que escribir cuarenta nombres a mano.
    for programa_malla, nombres in materias_por_programa().items():
        for nombre in nombres:
            if (programa_malla, normalizar(nombre)) in ya_en_base:
                continue
            _escribir_fila(ws, fila, [
                programa_malla, "", nombre, "", "", "SI", "",
            ], relleno=RELLENO_MALLA)
            fila += 1


def _hoja_previaturas(wb: Workbook, session: Session, nombres_programas: List[str]) -> None:
    columnas = [
        Columna("Programa*", 32, opciones=nombres_programas, obligatoria=True),
        Columna("Materia*\n(codigo o nombre)", 34, obligatoria=True),
        Columna("Materia previa*\n(codigo o nombre)", 34, obligatoria=True),
        Columna("¿Alcanza con aprobar\no tiene que exonerar?*", 22,
                opciones=TIPOS_PREVIATURA, obligatoria=True),
        Columna("Observaciones", 40),
    ]
    ws = _armar_hoja(
        wb, "4-Previaturas", columnas,
        nota="ESTA HOJA YA VIENE COMPLETA con la malla que nos pasaron. Solo revisala y "
             "corregí si algo cambio. Se lee: para cursar la MATERIA hay que tener la "
             "MATERIA PREVIA. APROBADA = le alcanza con aprobarla (por examen o "
             "exonerando). EXONERADA = solo sirve si la exonero. "
             "El Integrador de Analista Programador requiere TODAS las materias del "
             "programa: eso se carga solo, no hace falta una fila por cada una.",
    )

    previaturas = session.exec(select(Previatura)).all()
    materias = {m.id: m for m in session.exec(select(Materia)).all()}
    programas = {p.id: p.nombre for p in session.exec(select(Programa)).all()}

    fila = 3
    ya_en_base = set()
    for prev in previaturas:
        materia = materias.get(prev.materia_id)
        previa = materias.get(prev.materia_previa_id)
        if not materia or not previa:
            continue
        nombre_programa = programas.get(materia.programa_id, "")
        _escribir_fila(ws, fila, [
            nombre_programa, materia.codigo, previa.codigo,
            prev.tipo_requerido.value.upper(), "",
        ], relleno=RELLENO_PRECARGADO)
        ya_en_base.add((
            nombre_programa,
            normalizar(materia.nombre), normalizar(previa.nombre),
        ))
        fila += 1

    # La malla definida. Va por nombre porque los codigos todavia no existen: la
    # hoja acepta las dos cosas justamente por esto.
    for programa_malla, pares in previaturas_por_programa().items():
        for materia_nombre, previa_nombre in pares:
            clave = (
                programa_malla, normalizar(materia_nombre), normalizar(previa_nombre),
            )
            if clave in ya_en_base:
                continue
            _escribir_fila(ws, fila, [
                programa_malla, materia_nombre, previa_nombre, "APROBADA", "",
            ], relleno=RELLENO_MALLA)
            fila += 1


def _hoja_historial(wb: Workbook, nombres_programas: List[str]) -> None:
    columnas = [
        Columna("DOCUMENTO del alumno*", 20, obligatoria=True),
        Columna("Programa*", 32, opciones=nombres_programas, obligatoria=True),
        Columna("Materia*\n(codigo o nombre)", 30, obligatoria=True),
        Columna("Estado*", 16, opciones=ESTADOS_HISTORIAL, obligatoria=True),
        Columna("Nota", 10),
        Columna("Año", 10),
        Columna("Semestre\n(1 o 2)", 12),
        Columna("Observaciones", 40),
    ]
    _armar_hoja(
        wb, "5-Historial", columnas,
        nota="LA HOJA MAS IMPORTANTE. Una fila por alumno y materia que tenga algo. Si nunca "
             "la curso, no pongas la fila. Tiene que estar el historial COMPLETO de cada "
             "alumno, no solo lo de este año: el sistema controla previaturas y sin las "
             "materias viejas va a bloquear inscripciones que corresponden.",
    )


def _hoja_dictado(wb: Workbook, anio: int, nombres_programas: List[str]) -> None:
    columnas = [
        Columna("Programa*", 32, opciones=nombres_programas, obligatoria=True),
        Columna("Materia*\n(codigo o nombre)", 30, obligatoria=True),
        Columna("Año*", 10, obligatoria=True),
        Columna("Semestre*\n(1 o 2)", 12, obligatoria=True),
        Columna("DOCUMENTO del docente*", 20, obligatoria=True),
        Columna("Rol*", 14, opciones=ROLES_DOCENTE, obligatoria=True),
        Columna("Horario", 22),
        Columna("Salon", 14),
        Columna("Cupo maximo", 12),
        Columna("Observaciones", 40),
    ]
    _armar_hoja(
        wb, "6-Dictado actual", columnas,
        nota=f"Que se dicta este semestre y quien lo da. Si una materia tiene dos docentes, "
             f"dos filas con distinto rol. Cupo vacio = sin limite. "
             f"Empezá por {anio}; si queres cargar años anteriores para el historial "
             f"docente, agregalos con su año y semestre.",
    )


# ── Main ─────────────────────────────────────────────────────────────────────

def generar(salida: str, anio: int) -> str:
    wb = Workbook()
    wb.remove(wb.active)  # la hoja por defecto

    with get_db_session() as session:
        programas = list(session.exec(
            select(Programa).where(Programa.activo == True).order_by(Programa.nombre)
        ).all())
        nombres = [p.nombre for p in programas]

        # Sin programas no hay desplegables posibles; se deja abierto para que
        # los escriban a mano y el validador despues los reclame.
        if not nombres:
            nombres = []

        _hoja_leeme(wb, programas)
        _hoja_alumnos(wb, nombres)
        _hoja_docentes(wb)
        _hoja_plan(wb, session, nombres)
        _hoja_previaturas(wb, session, nombres)
        _hoja_historial(wb, nombres)
        _hoja_dictado(wb, anio, nombres)

    os.makedirs(os.path.dirname(os.path.abspath(salida)), exist_ok=True)
    wb.save(salida)
    return salida


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--salida",
        default=f"carga_inicial_portal_{datetime.now().strftime('%Y%m%d')}.xlsx",
        help="Archivo .xlsx a generar",
    )
    parser.add_argument(
        "--anio", type=int, default=datetime.now().year,
        help="Año lectivo de referencia para la hoja de dictado",
    )
    args = parser.parse_args()

    ruta = generar(args.salida, args.anio)
    print(f"Planilla generada: {os.path.abspath(ruta)}")
    print("\nHojas:")
    print("  LEEME               instrucciones para bedelia")
    print("  1-Alumnos           personas + a que carrera pertenecen")
    print("  2-Docentes          personas")
    print("  3-Plan de estudios  materias (precargado lo que ya existe)")
    print("  4-Previaturas       requisitos entre materias (precargado)")
    print("  5-Historial         alumno x materia x estado  <- la importante")
    print("  6-Dictado actual    quien dicta que este semestre")
    print("\nCuando la devuelvan:")
    print("  python -m v2.scripts.validar_planilla_migracion <archivo.xlsx>")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
