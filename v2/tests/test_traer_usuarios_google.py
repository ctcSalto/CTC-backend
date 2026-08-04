"""
Precarga de personas desde el directorio de Google.

Le ahorra a bedelia escribir 142 nombres y correos a mano. El correo importa
especialmente: es con lo que la persona entra al portal, asi que un error de
tipeo ahi la deja afuera.

Los tests van con respuestas de n8n armadas a mano, con las dos formas que se
vieron en produccion: con `orgUnitPath` (cuando el workflow no esta
simplificado) y sin el.
"""
import pytest
from openpyxl import load_workbook

from v2.models.enums import RolUsuario
from v2.scripts.traer_usuarios_google import (
    _usuarios_de, clasificar, escribir_alumnos, escribir_docentes,
)
from v2.scripts.validar_planilla_migracion import HOJAS


def cuenta(email, nombre, apellido, ou=None, suspendida=False):
    datos = {
        "primaryEmail": email,
        "name": {"givenName": nombre, "familyName": apellido},
        "suspended": suspendida,
    }
    if ou is not None:
        datos["orgUnitPath"] = ou
    return datos


class TestLecturaDeLaRespuesta:
    """n8n envuelve la respuesta de formas distintas segun el workflow."""

    def test_lista_directa(self):
        assert len(_usuarios_de([cuenta("a@x.uy", "A", "B")])) == 1

    def test_envuelta_en_users(self):
        respuesta = {"message": "ok", "users": [cuenta("a@x.uy", "A", "B")]}
        assert len(_usuarios_de(respuesta)) == 1

    def test_cada_item_envuelto_en_json(self):
        respuesta = {"users": [{"json": cuenta("a@x.uy", "A", "B")}]}
        usuarios = _usuarios_de(respuesta)
        assert usuarios[0]["primaryEmail"] == "a@x.uy"

    def test_respuesta_inesperada_no_rompe(self):
        assert _usuarios_de({"message": "sin datos"}) == []
        assert _usuarios_de(None) == []


class TestClasificacion:
    def test_reparte_por_unidad_organizativa(self):
        usuarios = [
            cuenta("a@x.uy", "Ana", "Perez", "/Alumnos"),
            cuenta("d@x.uy", "Luis", "Gomez", "/Equipo Docente"),
            cuenta("b@x.uy", "Eva", "Silva", "/Administración y Ventas"),
        ]
        por_rol = clasificar(usuarios)

        assert len(por_rol[RolUsuario.ESTUDIANTE]) == 1
        assert len(por_rol[RolUsuario.DOCENTE]) == 1
        assert len(por_rol[RolUsuario.ADMINISTRATIVO]) == 1

    def test_subniveles_de_ou(self):
        usuarios = [cuenta("a@x.uy", "Ana", "Perez", "/Alumnos/2026")]
        assert len(clasificar(usuarios)[RolUsuario.ESTUDIANTE]) == 1

    def test_sin_ou_queda_sin_clasificar(self):
        """
        getManyUsersGoogle tiene 'Simplify' activado y no devuelve orgUnitPath.
        No se puede asumir estudiante.
        """
        usuarios = [cuenta("a@x.uy", "Ana", "Perez")]
        por_rol = clasificar(usuarios)

        assert por_rol.get(RolUsuario.ESTUDIANTE) is None
        assert len(por_rol[None]) == 1

    def test_ou_no_mapeada_no_es_estudiante(self):
        """
        En el directorio real hay 8 casillas funcionales en /Gestión de Datos
        (soporte@, becas@, servidor@) y 4 cuentas en la raiz. Antes del arreglo
        de ou_to_rol, las 12 entraban como alumnos.
        """
        usuarios = [
            cuenta("soporte@x.uy", "Soporte", "CTC", "/Gestión de Datos"),
            cuenta("admisiones@x.uy", "Admisiones", "CTC", "/"),
        ]
        por_rol = clasificar(usuarios)

        assert por_rol.get(RolUsuario.ESTUDIANTE) is None
        assert len(por_rol[None]) == 2


class TestResolucionDeOUPorUsuario:
    """
    Plan B cuando el lote no trae la OU: se consulta de a una persona contra
    el workflow getGoogleUO.
    """

    def test_completa_las_que_faltan(self, monkeypatch):
        from v2.scripts import traer_usuarios_google as modulo
        import v2.auth.n8n_ou_client as cliente

        monkeypatch.setattr(
            cliente.n8n_ou_client, "get_user_ou",
            lambda email: "/Alumnos" if "alumno" in email else "/Equipo Docente",
        )

        usuarios = [
            cuenta("alumno@x.uy", "Ana", "Perez"),
            cuenta("profe@x.uy", "Luis", "Gomez"),
        ]
        resueltas = modulo.completar_ou(usuarios)

        assert resueltas == 2
        assert usuarios[0]["orgUnitPath"] == "/Alumnos"
        assert usuarios[1]["orgUnitPath"] == "/Equipo Docente"

    def test_no_reconsulta_las_que_ya_tienen_ou(self, monkeypatch):
        from v2.scripts import traer_usuarios_google as modulo
        import v2.auth.n8n_ou_client as cliente

        llamadas = []
        monkeypatch.setattr(
            cliente.n8n_ou_client, "get_user_ou",
            lambda email: llamadas.append(email) or "/Alumnos",
        )

        usuarios = [cuenta("a@x.uy", "Ana", "Perez", "/Equipo Docente")]
        modulo.completar_ou(usuarios)

        assert llamadas == []
        assert usuarios[0]["orgUnitPath"] == "/Equipo Docente"

    def test_si_la_consulta_falla_no_inventa(self, monkeypatch):
        from v2.scripts import traer_usuarios_google as modulo
        import v2.auth.n8n_ou_client as cliente

        monkeypatch.setattr(cliente.n8n_ou_client, "get_user_ou", lambda email: None)

        usuarios = [cuenta("a@x.uy", "Ana", "Perez")]
        assert modulo.completar_ou(usuarios) == 0
        assert usuarios[0].get("orgUnitPath") is None


class TestEscrituraEnLaPlanilla:
    @pytest.fixture(name="planilla")
    def fixture_planilla(self, tmp_path, session, programa, materias_con_previaturas):
        from contextlib import contextmanager
        import v2.scripts.generar_planilla_migracion as modulo
        from v2.scripts.generar_planilla_migracion import generar

        @contextmanager
        def sesion_de_test():
            yield session

        original = modulo.get_db_session
        modulo.get_db_session = sesion_de_test
        try:
            ruta = str(tmp_path / "planilla.xlsx")
            generar(ruta, anio=2026)
        finally:
            modulo.get_db_session = original
        return ruta

    def test_escribe_apellido_nombre_y_correo(self, planilla):
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]

        escritos = escribir_alumnos(
            ws, [cuenta("ana.perez@ctcsalto.edu.uy", "Ana", "Perez", "/Alumnos")],
            sin_ou=False,
        )

        assert escritos == 1
        assert ws.cell(row=3, column=2).value == "Perez"
        assert ws.cell(row=3, column=3).value == "Ana"
        assert ws.cell(row=3, column=4).value == "ana.perez@ctcsalto.edu.uy"

    def test_no_inventa_la_cedula(self, planilla):
        """Google no la tiene: la columna queda vacia para que la complete bedelia."""
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]

        escribir_alumnos(
            ws, [cuenta("ana.perez@ctcsalto.edu.uy", "Ana", "Perez", "/Alumnos")],
            sin_ou=False,
        )

        assert ws.cell(row=3, column=1).value is None

    def test_correr_dos_veces_no_duplica(self, planilla):
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]
        gente = [cuenta("ana.perez@ctcsalto.edu.uy", "Ana", "Perez", "/Alumnos")]

        assert escribir_alumnos(ws, gente, sin_ou=False) == 1
        assert escribir_alumnos(ws, gente, sin_ou=False) == 0

    def test_no_pisa_lo_que_ya_escribio_bedelia(self, planilla):
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]
        ws.cell(row=3, column=1, value="41234567")
        ws.cell(row=3, column=2, value="Perez")
        ws.cell(row=3, column=4, value="ana.perez@ctcsalto.edu.uy")

        escritos = escribir_alumnos(
            ws, [cuenta("ana.perez@ctcsalto.edu.uy", "Ana", "Perez", "/Alumnos")],
            sin_ou=False,
        )

        assert escritos == 0
        assert ws.cell(row=3, column=1).value == "41234567"

    def test_marca_las_filas_sin_ou(self, planilla):
        """Si no se sabe el rol, tiene que quedar visible en la planilla."""
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]

        escribir_alumnos(
            ws, [cuenta("ana.perez@ctcsalto.edu.uy", "Ana", "Perez")], sin_ou=True,
        )

        assert "VERIFICAR ROL" in ws.cell(row=3, column=11).value

    def test_marca_las_cuentas_suspendidas(self, planilla):
        wb = load_workbook(planilla)
        ws = wb[HOJAS["alumnos"]]

        escribir_alumnos(
            ws,
            [cuenta("ex@ctcsalto.edu.uy", "Ex", "Alumno", "/Alumnos", suspendida=True)],
            sin_ou=False,
        )

        assert "suspendida" in ws.cell(row=3, column=11).value

    def test_docente_suspendido_queda_inactivo(self, planilla):
        """La hoja de docentes tiene su propia columna de activo."""
        wb = load_workbook(planilla)
        ws = wb[HOJAS["docentes"]]

        escribir_docentes(
            ws,
            [cuenta("ex@ctcsalto.edu.uy", "Ex", "Docente", "/Equipo Docente",
                    suspendida=True)],
        )

        assert ws.cell(row=3, column=7).value == "NO"

    def test_docente_activo(self, planilla):
        wb = load_workbook(planilla)
        ws = wb[HOJAS["docentes"]]

        escribir_docentes(
            ws, [cuenta("d@ctcsalto.edu.uy", "Luis", "Gomez", "/Equipo Docente")],
        )

        assert ws.cell(row=3, column=7).value == "SI"


class TestValidacionDeLoPrecargado:
    """
    Lo que trae Google no alcanza para importar: falta la cedula. El validador
    tiene que decirlo una vez, no ciento cuarenta veces.
    """

    def test_las_cedulas_faltantes_se_informan_juntas(
        self, tmp_path, session, programa, materias_con_previaturas
    ):
        from contextlib import contextmanager
        import v2.scripts.generar_planilla_migracion as modulo
        from v2.scripts.generar_planilla_migracion import generar
        from v2.scripts.validar_planilla_migracion import Validador

        @contextmanager
        def sesion_de_test():
            yield session

        original = modulo.get_db_session
        modulo.get_db_session = sesion_de_test
        try:
            ruta = str(tmp_path / "planilla.xlsx")
            generar(ruta, anio=2026)
        finally:
            modulo.get_db_session = original

        wb = load_workbook(ruta)
        gente = [
            cuenta(f"persona{i}@ctcsalto.edu.uy", f"N{i}", f"A{i}", "/Alumnos")
            for i in range(30)
        ]
        escribir_alumnos(wb[HOJAS["alumnos"]], gente, sin_ou=False)
        wb.save(ruta)

        problemas = Validador(ruta).correr()
        sin_cedula = [p for p in problemas if "sin cedula" in p.mensaje]

        assert len(sin_cedula) == 1
        assert "30 alumnos" in sin_cedula[0].mensaje
