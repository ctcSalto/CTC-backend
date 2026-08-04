"""
Planilla de carga inicial: generacion y validacion.

Estos scripts se corren un puñado de veces y despues no se tocan mas, pero
deciden que datos entran al sistema el dia del despliegue. Si el validador deja
pasar algo, el error aparece en produccion con alumnos reales adentro.

El chequeo de cadenas incompletas es el que mas importa: es el que anticipa los
"no me deja inscribirme" del primer dia.
"""
import pytest
from openpyxl import load_workbook

from v2.scripts.generar_planilla_migracion import generar
from v2.scripts.malla_inicial import normalizar
from v2.scripts.validar_planilla_migracion import HOJAS, Validador


@pytest.fixture(name="planilla_vacia")
def fixture_planilla_vacia(tmp_path, session, programa, materias_con_previaturas):
    """
    Genera la planilla contra la base de test.

    `generar` abre su propia sesion con get_db_session, asi que se le pasa la
    del fixture por monkeypatch en el propio modulo.
    """
    from contextlib import contextmanager
    import v2.scripts.generar_planilla_migracion as modulo

    @contextmanager
    def sesion_de_test():
        yield session

    original = modulo.get_db_session
    modulo.get_db_session = sesion_de_test
    try:
        ruta = str(tmp_path / "planilla.xlsx")
        generar(ruta, anio=2026)
    finally:
        modulo.get_db_session = original
    return completar_malla(ruta)


def completar_malla(ruta):
    """
    Rellena semestre y creditos de las filas precargadas desde la malla.

    Es lo primero que hace bedelia: la planilla llega con ~40 materias que solo
    tienen el nombre. Sin esto, todo test que espere cero errores choca contra
    esas filas incompletas, que es justamente lo que el validador debe pedir.
    """
    wb = load_workbook(ruta)
    ws = wb[HOJAS["plan"]]
    for fila in range(3, ws.max_row + 1):
        if ws.cell(row=fila, column=3).value is None:
            continue
        if ws.cell(row=fila, column=4).value is None:
            ws.cell(row=fila, column=4, value=1)
        if ws.cell(row=fila, column=5).value is None:
            ws.cell(row=fila, column=5, value=10)
    wb.save(ruta)
    return ruta


def escribir(ruta, hoja, filas, desde=3):
    """Escribe filas en una hoja, desde la primera fila de datos."""
    wb = load_workbook(ruta)
    ws = wb[hoja]
    for indice, fila in enumerate(filas, start=desde):
        for columna, valor in enumerate(fila, start=1):
            ws.cell(row=indice, column=columna, value=valor)
    wb.save(ruta)
    return ruta


def problemas_de(ruta):
    validador = Validador(ruta)
    problemas = validador.correr()
    return (
        [p for p in problemas if p.nivel == "ERROR"],
        [p for p in problemas if p.nivel == "AVISO"],
    )


class TestGeneracion:
    def test_tiene_todas_las_hojas(self, planilla_vacia):
        wb = load_workbook(planilla_vacia)
        for nombre in HOJAS.values():
            assert nombre in wb.sheetnames
        assert "LEEME" in wb.sheetnames

    def test_precarga_las_materias_que_ya_existen(self, planilla_vacia):
        """Bedelia no tiene que tipear lo que el sistema ya sabe."""
        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["plan"]]
        codigos = {
            ws.cell(row=fila, column=2).value
            for fila in range(3, ws.max_row + 1)
        }
        assert {"P1_T", "P2_T", "P3_T"} <= codigos

    def test_precarga_las_previaturas(self, planilla_vacia):
        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["previaturas"]]
        pares = {
            (ws.cell(row=fila, column=2).value, ws.cell(row=fila, column=3).value)
            for fila in range(3, ws.max_row + 1)
        }
        assert ("P2_T", "P1_T") in pares
        assert ("P3_T", "P2_T") in pares

    def test_el_estado_del_historial_es_desplegable(self, planilla_vacia):
        """Sin lista cerrada, la columna vuelve con diez formas de decir aprobado."""
        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["historial"]]
        formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
        assert any("APROBADA" in f and "A_EXAMEN" in f for f in formulas)


class TestMallaPrecargada:
    """
    Las 44 previaturas ya estaban definidas en el repo desde mayo, en la
    migracion seed_previaturas, pero esa migracion no inserto nada: resolvia las
    materias por nombre exacto contra una tabla vacia. Se precargan en la
    planilla para no pedirle a bedelia que las tipee de nuevo.
    """

    def test_las_materias_de_la_malla_estan_en_el_plan(self, planilla_vacia):
        from v2.scripts.malla_inicial import materias_por_programa

        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["plan"]]
        nombres = {
            normalizar(str(ws.cell(row=fila, column=3).value or ""))
            for fila in range(3, ws.max_row + 1)
        }

        esperadas = materias_por_programa()["Analista Programador"]
        faltantes = [n for n in esperadas if normalizar(n) not in nombres]
        assert faltantes == [], faltantes

    def test_las_previaturas_de_la_malla_estan_cargadas(self, planilla_vacia):
        from v2.scripts.malla_inicial import previaturas_por_programa

        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["previaturas"]]
        pares = {
            (normalizar(str(ws.cell(row=fila, column=2).value or "")),
             normalizar(str(ws.cell(row=fila, column=3).value or "")))
            for fila in range(3, ws.max_row + 1)
        }

        for programa, esperados in previaturas_por_programa().items():
            for materia, previa in esperados:
                assert (normalizar(materia), normalizar(previa)) in pares, \
                    f"falta {materia} -> {previa} de {programa}"

    def test_la_planilla_recien_generada_pide_semestre_y_creditos(
        self, tmp_path, session, programa, materias_con_previaturas
    ):
        """Sin completar, el validador tiene que reclamar las filas de la malla."""
        from contextlib import contextmanager
        import v2.scripts.generar_planilla_migracion as modulo

        @contextmanager
        def sesion_de_test():
            yield session

        original = modulo.get_db_session
        modulo.get_db_session = sesion_de_test
        try:
            ruta = str(tmp_path / "sin_completar.xlsx")
            generar(ruta, anio=2026)
        finally:
            modulo.get_db_session = original

        errores, _ = problemas_de(ruta)
        reclamos = [e for e in errores if "sin completar" in str(e)]
        assert len(reclamos) == 1, [str(e) for e in errores]
        assert "semestre" in str(reclamos[0])

    def test_previaturas_por_nombre_se_resuelven(self, planilla_vacia, programa):
        """
        La malla viene con nombres, no con codigos: los codigos se decidieron
        despues. La hoja tiene que aceptar las dos formas.
        """
        errores, _ = problemas_de(planilla_vacia)
        assert not [e for e in errores if "Previaturas" in str(e)], \
            [str(e) for e in errores]

    def test_historial_acepta_el_nombre_de_la_materia(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [
            ("41234567", "Perez", "Ana", None, None, None, None,
             programa.nombre, 2023, "ACTIVA", None)
        ])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "Programacion 1", "APROBADA", 78, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert errores == [], [str(e) for e in errores]


class TestValidacion:
    def _alumno(self, documento="41234567", programa="Ingenieria de Test"):
        return (documento, "Perez", "Ana", None, None, None, None,
                programa, 2023, "ACTIVA", None)

    def test_planilla_bien_llena_no_da_errores(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "APROBADA", 78, 2024, 1, None),
            ("41234567", programa.nombre, "P2_T", "CURSANDO", None, 2026, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert errores == [], [str(e) for e in errores]

    def test_documento_con_puntos_se_normaliza(self, planilla_vacia, programa):
        """Bedelia escribe 4.123.456-7 y el sistema tiene que entenderlo."""
        escribir(planilla_vacia, HOJAS["alumnos"], [
            self._alumno("4.123.456-7", programa.nombre)
        ])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "APROBADA", 78, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert errores == [], [str(e) for e in errores]

    def test_alumno_del_historial_sin_ficha(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("99999999", programa.nombre, "P1_T", "APROBADA", 78, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("99999999" in str(e) for e in errores)

    def test_materia_inexistente(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "NO_EXISTE", "APROBADA", 78, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("NO_EXISTE" in str(e) for e in errores)

    def test_una_sola_fila_por_alumno_y_materia(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "APROBADA", 78, 2024, 1, None),
            ("41234567", programa.nombre, "P1_T", "EXONERADA", 90, 2025, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("una fila para 'P1_T'" in str(e) for e in errores)

    def test_estado_fuera_del_vocabulario(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "aprobo", 78, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("APROBADA" in str(e) for e in errores)

    def test_nota_fuera_de_rango(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa=programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "APROBADA", 150, 2024, 1, None),
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("150" in str(e) for e in errores)

    def test_programa_que_no_esta_en_el_plan(self, planilla_vacia):
        escribir(planilla_vacia, HOJAS["alumnos"], [
            self._alumno(programa="Carrera Que No Existe")
        ])

        errores, _ = problemas_de(planilla_vacia)
        assert any("Carrera Que No Existe" in str(e) for e in errores)

    def test_ciclo_de_previaturas(self, planilla_vacia, programa):
        """P1 requiere P3 cierra el circulo con las precargadas."""
        wb = load_workbook(planilla_vacia)
        ws = wb[HOJAS["previaturas"]]
        fila = ws.max_row + 1
        for columna, valor in enumerate(
            (programa.nombre, "P1_T", "P3_T", "APROBADA", None), start=1
        ):
            ws.cell(row=fila, column=columna, value=valor)
        wb.save(planilla_vacia)

        errores, _ = problemas_de(planilla_vacia)
        assert any("Ciclo de previaturas" in str(e) for e in errores)


class TestCadenasIncompletas:
    """
    El chequeo que anticipa los reclamos del dia uno: si falta el historial
    viejo, el portal bloquea inscripciones que en realidad corresponden.
    """

    def _alumno(self, programa):
        return ("41234567", "Perez", "Ana", None, None, None, None,
                programa, 2023, "ACTIVA", None)

    def test_avisa_si_falta_la_previatura(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa.nombre)])
        # P2 aprobada, pero P1 (su previatura) no figura
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P2_T", "APROBADA", 78, 2024, 1, None),
        ])

        errores, avisos = problemas_de(planilla_vacia)
        assert errores == [], [str(e) for e in errores]
        assert any("P1_T" in str(a) and "no figura" in str(a) for a in avisos)

    def test_avisa_si_la_previatura_esta_pero_sin_aprobar(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "RECURSA", 40, 2023, 1, None),
            ("41234567", programa.nombre, "P2_T", "APROBADA", 78, 2024, 1, None),
        ])

        errores, avisos = problemas_de(planilla_vacia)
        assert errores == []
        assert any("P1_T" in str(a) and "RECURSA" in str(a) for a in avisos)

    def test_cadena_completa_no_avisa(self, planilla_vacia, programa):
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "APROBADA", 78, 2023, 1, None),
            ("41234567", programa.nombre, "P2_T", "EXONERADA", 90, 2024, 1, None),
            ("41234567", programa.nombre, "P3_T", "CURSANDO", None, 2026, 1, None),
        ])

        errores, avisos = problemas_de(planilla_vacia)
        assert errores == []
        assert not [a for a in avisos if "previatura" in str(a)], [str(a) for a in avisos]

    def test_exonerada_cuenta_como_tenida(self, planilla_vacia, programa):
        """Exonerar es aprobar sin examen: no puede pedir la previa de nuevo."""
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "EXONERADA", 95, 2023, 1, None),
            ("41234567", programa.nombre, "P2_T", "APROBADA", 78, 2024, 1, None),
        ])

        _, avisos = problemas_de(planilla_vacia)
        assert not [a for a in avisos if "previatura" in str(a)]

    def test_a_examen_no_cuenta_como_tenida(self, planilla_vacia, programa):
        """Tener derecho a examen no es tener la materia."""
        escribir(planilla_vacia, HOJAS["alumnos"], [self._alumno(programa.nombre)])
        escribir(planilla_vacia, HOJAS["historial"], [
            ("41234567", programa.nombre, "P1_T", "A_EXAMEN", 72, 2023, 1, None),
            ("41234567", programa.nombre, "P2_T", "APROBADA", 78, 2024, 1, None),
        ])

        _, avisos = problemas_de(planilla_vacia)
        assert any("P1_T" in str(a) and "A_EXAMEN" in str(a) for a in avisos)
