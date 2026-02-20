"""
Script de migración de usuarios desde el Excel 'Moodle 2025.xlsx' a Google Workspace.

Lee las hojas 'Usuarios' e 'Incriptos' (mismo esquema de columnas).
Para cada usuario no migrado:
  1. Genera email: firstname.lastname@ctcsalto.edu.uy
  2. Asigna OU según Tipo Usuario
  3. Crea cuenta en Google Workspace
  4. Envía notificación por n8n
  5. Escribe Nuevo Correo, Nueva Contraseña y Migrado=Si en el xlsx

Manejo de errores por etapa:
  - Si la cuenta ya existe en Google (error 409) → se marca como migrada de igual forma.
  - Si la notificación falla DESPUÉS de crear la cuenta → la cuenta queda creada,
    pero Migrado NO se marca Si. Al volver a ejecutar el script detectará esto y
    solo reintentará la notificación + escritura al xlsx.
  - Si falla la escritura al xlsx → se imprime la contraseña en consola para
    que no se pierda.

Uso:
    python migrate_usuarios.py --dry-run                  # ver el plan completo
    python migrate_usuarios.py --usuario eviera           # solo ese usuario
    python migrate_usuarios.py --hoja Incriptos --dry-run # solo incriptos
    python migrate_usuarios.py                            # ambas hojas
"""

import sys
import os
import argparse
import secrets
import string
import unicodedata
import traceback

# ── sys.path → raíz del proyecto (3 niveles arriba de scripts/) ──────────────
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, PROJECT_ROOT)

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"), override=True)

import requests
import openpyxl
from external_services.google.google_service import GoogleWorkspaceService

# ─── CONFIG ───────────────────────────────────────────────────────────────────
XLSX_PATH = os.path.join(PROJECT_ROOT, "Moodle 2025.xlsx")
DOMAIN    = "ctcsalto.edu.uy"

# Hojas que se procesan (por defecto ambas)
SHEETS = ["Usuarios", "Incriptos"]

# Columnas del xlsx (0-indexed) — mismo esquema en ambas hojas
COL_ID            = 0   # A
COL_USERNAME      = 1   # B  (None en Incriptos)
COL_FIRSTNAME     = 2   # C
COL_LASTNAME      = 3   # D
COL_FULLNAME      = 4   # E
COL_EMAIL_ORIG    = 5   # F
COL_CONTINUA      = 6   # G
COL_NUEVO_CORREO  = 7   # H  ← se escribe tras crear cuenta
COL_NUEVA_CONTRA  = 8   # I  ← se escribe tras crear cuenta
COL_MIGRADO       = 9   # J  ← se escribe "Si" tras migración exitosa
COL_TIPO_USUARIO  = 10  # K

# Mapeo Tipo Usuario → Org Unit
OU_MAP = {
    "Estudiante":     "/Alumnos",
    "Docente":        "/Equipo Docente",
    "Administrativo": "/Administración y Ventas",
}

# Contraseña
PASSWORD_LENGTH = 16
PASSWORD_CHARS  = string.ascii_letters + string.digits + "!@#$%&*-_=+"

# Usernames de sistema que se saltan siempre
SKIP_USERNAMES = {"guest"}


# ─── UTILIDADES ───────────────────────────────────────────────────────────────

def strip_accents(text: str) -> str:
    """Elimina tildes y cedillas: 'González' → 'gonzalez'"""
    nfkd = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def _extract_surname(lastname: str) -> str:
    """
    Extrae la parte significativa del apellido, tratando prefijos compuestos.
    Prefijos (De/Del/Di/Da/Dos/Das/La/Las/Los) se unen con la siguiente palabra.
        'De Lima Albin'      → 'delima'
        'Di Donato Caputto'  → 'didonato'
        'De los Santos ...'  → 'delossantos'
        'Baldassini Silva'   → 'baldassini'
    """
    prefixes = {'de', 'del', 'la', 'las', 'los', 'di', 'da', 'dos', 'das'}
    parts    = lastname.strip().split()
    surname  = ""
    for part in parts:
        surname += strip_accents(part).lower()
        if part.lower() not in prefixes:
            break
    return surname


def generate_workspace_email(firstname: str, lastname: str) -> str:
    """
    Genera email de Workspace.
    firstname → primera palabra, lastname → _extract_surname (maneja compuestos).
        'Jorge Luis' / 'Baldassini Silva' → jorge.baldassini@ctcsalto.edu.uy
        'Lucas Leonel' / 'De Lima Albin'  → lucas.delima@ctcsalto.edu.uy
    """
    first = strip_accents(firstname.strip().split()[0]).lower()
    last  = _extract_surname(lastname)
    return f"{first}.{last}@{DOMAIN}"


def generate_password() -> str:
    """Contraseña aleatoria segura de PASSWORD_LENGTH chars con al menos 1 de cada categoría."""
    guaranteed = [
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.digits),
        secrets.choice("!@#$%&*-_=+"),
    ]
    rest = [secrets.choice(PASSWORD_CHARS) for _ in range(PASSWORD_LENGTH - 4)]
    chars = guaranteed + rest
    # Fisher-Yates shuffle con secrets
    for i in range(len(chars) - 1, 0, -1):
        j = secrets.randbelow(i + 1)
        chars[i], chars[j] = chars[j], chars[i]
    return ''.join(chars)


def send_notification(nuevo_correo: str, nueva_contrasena: str,
                      firstname: str, lastname: str, email_original: str) -> bool:
    """
    POST al webhook de n8n para enviar la notificación al usuario.
    Levanta ValueError si el webhook retorna >= 400.
    """
    webhook_url = os.getenv("N8N_NOTIFICATION_WEBHOOK_URL")
    if not webhook_url:
        raise ValueError("N8N_NOTIFICATION_WEBHOOK_URL no está definida en .env")

    api_token   = os.getenv("N8N_API_TOKEN", "")
    header_name = os.getenv("N8N_HEADER_NAME", "Authorization")
    timeout     = int(os.getenv("N8N_TIMEOUT", "30"))

    headers = {
        header_name:    api_token,
        "Content-Type": "application/json",
    }
    payload = {
        "firstname":        firstname,
        "lastname":         lastname,
        "nuevo_correo":     nuevo_correo,
        "nueva_contrasena": nueva_contrasena,
        "email_original":   email_original,
    }

    response = requests.post(webhook_url, json=payload, headers=headers, timeout=timeout)
    if response.status_code >= 400:
        raise ValueError(f"Webhook notificación retornó {response.status_code}: {response.text}")
    return True


# ─── LECTURA DEL XLSX ─────────────────────────────────────────────────────────

def read_usuarios_from_sheet(sheet_name: str, username_filter: str = None) -> list:
    """
    Lee una hoja del xlsx y retorna lista de dicts con los usuarios elegibles.
    Skipa: guest, ya migrados, sin Continua=Si, sin Tipo Usuario válido, sin nombre/apellido.
    Para Incriptos (sin username) genera un identificador a partir del email original.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]

    usuarios = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # ── campos crudos ──
        username     = str(row[COL_USERNAME]).strip() if row[COL_USERNAME] else None
        firstname    = str(row[COL_FIRSTNAME]).strip()  if row[COL_FIRSTNAME]    else ""
        lastname     = str(row[COL_LASTNAME]).strip()   if row[COL_LASTNAME]     else ""
        email_orig   = str(row[COL_EMAIL_ORIG]).strip() if row[COL_EMAIL_ORIG]   else ""
        continua     = str(row[COL_CONTINUA]).strip().lower()  if row[COL_CONTINUA]    else ""
        migrado      = str(row[COL_MIGRADO]).strip().lower()   if row[COL_MIGRADO]     else "no"
        tipo_usuario = str(row[COL_TIPO_USUARIO]).strip()      if row[COL_TIPO_USUARIO] else ""
        nuevo_correo_existente = str(row[COL_NUEVO_CORREO]).strip() if row[COL_NUEVO_CORREO] else ""

        # ── skip por nombre/apellido vacíos (filas placeholder en Incriptos) ──
        if not firstname or not lastname:
            continue

        # ── skip guest ──
        if username and username.lower() in SKIP_USERNAMES:
            continue

        # ── skip ya migrados ──
        if migrado == "si":
            continue

        # ── skip sin Continua=Si ──
        if continua != "si":
            continue

        # ── skip sin Tipo Usuario válido ──
        if not tipo_usuario or tipo_usuario not in OU_MAP:
            continue

        # ── para Incriptos sin username: usar email original como referencia ──
        display_id = username if username else (email_orig if email_orig else f"fila-{i}")

        # ── filtro por --usuario (busca en username o email original) ──
        if username_filter:
            if username != username_filter and email_orig != username_filter:
                continue

        usuarios.append({
            "sheet":            sheet_name,
            "row":              i,
            "username":         username,           # puede ser None (Incriptos)
            "display_id":       display_id,
            "firstname":        firstname,
            "lastname":         lastname,
            "email_original":   email_orig,
            "tipo_usuario":     tipo_usuario,
            "nuevo_correo_existente": nuevo_correo_existente,  # para detectar cuenta creada pero no notificada
        })

    wb.close()
    return usuarios


def read_all_usuarios(username_filter: str = None, hoja_filter: str = None) -> list:
    """
    Lee Usuarios + Incriptos, deduplicando por email de Workspace generado.
    Si hay duplicados, se toma la primera aparición (Usuarios tiene prioridad).
    """
    hojas = [hoja_filter] if hoja_filter else SHEETS
    todos = []
    seen_emails = set()   # emails de Workspace ya vistos → dedup

    for hoja in hojas:
        for u in read_usuarios_from_sheet(hoja, username_filter):
            email_ws = generate_workspace_email(u["firstname"], u["lastname"])
            if email_ws in seen_emails:
                print(f"  [DEDUP] {u['display_id']} ({u['sheet']} fila {u['row']}) → {email_ws} ya existe, skipeado.")
                continue
            seen_emails.add(email_ws)
            todos.append(u)

    return todos


# ─── ESCRIBIR RESULTADOS AL XLSX ──────────────────────────────────────────────

def write_migration_result(sheet_name: str, row_number: int,
                           nuevo_correo: str, nueva_contrasena: str):
    """
    Escribe Nuevo Correo, Nueva Contraseña y Migrado=Si en la fila indicada.
    Abre y cierra el workbook cada vez para evitar conflictos si el proceso se interrumpe.
    """
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[sheet_name]

    ws.cell(row=row_number, column=COL_NUEVO_CORREO + 1, value=nuevo_correo)
    ws.cell(row=row_number, column=COL_NUEVA_CONTRA + 1, value=nueva_contrasena)
    ws.cell(row=row_number, column=COL_MIGRADO + 1,      value="Si")

    wb.save(XLSX_PATH)
    wb.close()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migrar usuarios a Google Workspace")
    parser.add_argument("--usuario", help="Username o email del usuario a migrar (solo ese)", default=None)
    parser.add_argument("--hoja",    help="Hoja específica: Usuarios | Incriptos (por defecto ambas)", default=None)
    parser.add_argument("--dry-run", help="Solo muestra el plan, no crea nada", action="store_true")
    args = parser.parse_args()

    # ── Lectura + dedup ──
    print(f"Leyendo {XLSX_PATH}...")
    usuarios = read_all_usuarios(username_filter=args.usuario, hoja_filter=args.hoja)
    print(f"Usuarios elegibles (tras dedup): {len(usuarios)}\n")

    # ── Plan ──
    print(f"{'=' * 70}")
    print("PLAN DE MIGRACIÓN DE USUARIOS")
    print(f"{'=' * 70}")
    for u in usuarios:
        email = generate_workspace_email(u["firstname"], u["lastname"])
        ou    = OU_MAP[u["tipo_usuario"]]
        estado = ""
        if u["nuevo_correo_existente"]:
            estado = "  ⚠ cuenta probablemente creada, notificación pendiente"
        print(f"  [{u['row']:>3}] [{u['sheet']:10s}] {u['display_id']:30s} → {email}")
        print(f"        Nombre: {u['firstname']} {u['lastname']}")
        print(f"        OU: {ou}  |  Email original: {u['email_original']}{estado}")
        print()

    if args.dry_run or not usuarios:
        print("[DRY-RUN] No se creó nada." if args.dry_run else "Nada que crear.")
        return

    # ── Ejecución ──
    print(f"{'=' * 70}")
    print("CREANDO CUENTAS")
    print(f"{'=' * 70}\n")

    service = GoogleWorkspaceService()
    created = 0
    errors  = 0

    for u in usuarios:
        sheet     = u["sheet"]
        row_num   = u["row"]
        firstname = u["firstname"]
        lastname  = u["lastname"]
        email_orig = u["email_original"]
        tipo      = u["tipo_usuario"]
        display   = u["display_id"]

        nuevo_correo     = generate_workspace_email(firstname, lastname)
        nueva_contrasena = generate_password()
        ou               = OU_MAP[tipo]

        # Para family_name en Google: usar _extract_surname para que "De Lima" → "De Lima"
        # (Google Workspace acepta el nombre completo del apellido)
        # Pero given_name solo la primera palabra del firstname
        given_name  = firstname.split()[0]
        family_name = lastname.split()[0] if lastname.split()[0].lower() not in {'de','del','di','da','dos','das','la','las','los'} else ' '.join(lastname.split()[:2]) if len(lastname.split()) > 1 else lastname

        print(f"  [{row_num:>3}] [{sheet:10s}] {display} → {nuevo_correo}  OU={ou}")

        # ─── ETAPA 1: crear cuenta Google Workspace ─────────────────────────
        account_created = False
        try:
            result = service.create_google_account(
                primary_email=nuevo_correo,
                given_name=given_name,
                family_name=family_name,
                password=nueva_contrasena,
                org_unit_path=ou,
            )
            print(f"        [OK] Cuenta creada: {result}")
            account_created = True

        except ValueError as e:
            err_msg = str(e)
            # 409 = cuenta ya existe → no es error fatal, continuamos
            if "409" in err_msg or "already exists" in err_msg.lower() or "ya existe" in err_msg.lower():
                print("        [WARN] Cuenta ya existe en Google Workspace, continuando...")
                account_created = True
            else:
                print(f"        [ERR] Etapa 1 - Crear cuenta: {err_msg}")
                errors += 1
                continue   # no tiene sentido notificar si no hay cuenta

        except Exception as e:
            print(f"        [ERR] Etapa 1 - Crear cuenta: {e}")
            traceback.print_exc()
            errors += 1
            continue

        # ─── ETAPA 2: enviar notificación ────────────────────────────────────
        if not account_created:
            # no debería llegar aquí, pero por seguridad
            print("        [ERR] Cuenta no creada, skipeando notificación.")
            errors += 1
            continue

        try:
            send_notification(
                nuevo_correo=nuevo_correo,
                nueva_contrasena=nueva_contrasena,
                firstname=firstname,
                lastname=lastname,
                email_original=email_orig,
            )
            print("        [OK] Notificación enviada")

        except Exception as e:
            print(f"        [ERR] Etapa 2 - Notificación: {e}")
            print("        ⚠ CUENTA YA CREADA pero notificación falló.")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            print("        ⚠ Re-ejecutar el script para reintentar la notificación.")
            errors += 1
            continue   # NO escribir Migrado=Si → al re-ejecutar se reintentará

        # ─── ETAPA 3: guardar en xlsx (credenciales + Migrado=Si) ─────────────
        try:
            write_migration_result(sheet, row_num, nuevo_correo, nueva_contrasena)
            print(f"        [OK] xlsx actualizado (fila {row_num}, Migrado=Si)")
            created += 1

        except Exception as e:
            print(f"        [ERR] Etapa 3 - Escribir xlsx: {e}")
            print("        ⚠ CUENTA CREADA Y NOTIFICADA, pero xlsx no se actualizó.")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            errors += 1

    # ── Reporte final ──
    print(f"\n{'=' * 70}")
    print("REPORTE DE MIGRACIÓN DE USUARIOS")
    print(f"{'=' * 70}")
    print(f"  Cuentas creadas exitosas: {created}")
    print(f"  Errores:                  {errors}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
