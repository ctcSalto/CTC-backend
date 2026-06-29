"""
Script de migración: Cursos Marzo 2026.

Lee el xlsx "Creación mail institucional y acceso moodle - Cursos 2026 _ Marzo.xlsx"
con estructura: Estudiantes | Mail | Curso (filas 4-16) y Docente | Mail | Curso (fila 22).

Para cada persona:
  1. Parsea primer nombre + primer apellido (partículas como "de" se pegan al apellido)
  2. Genera email: nombre.apellido@ctcsalto.edu.uy
  3. Crea cuenta en Google Workspace (Alumnos o Equipo Docente)
  4. Envía notificación por n8n al email personal
  5. Escribe Nuevo Correo, Nueva Contraseña y Migrado=Si en el xlsx

Uso:
    python migrate_cursos_marzo_2026.py --dry-run     # ver plan sin ejecutar
    python migrate_cursos_marzo_2026.py               # todos los pendientes
"""

import sys
import os
import argparse
import traceback

PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, PROJECT_ROOT)

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"), override=True)

import openpyxl
from external_services.google.google_service import GoogleWorkspaceService

from migrate_usuarios import (
    strip_accents,
    generate_password,
    send_notification,
    DOMAIN,
)

# ─── CONFIG ───────────────────────────────────────────────────────────────────
XLSX_PATH = os.path.join(PROJECT_ROOT, "Creación mail institucional y acceso moodle - Cursos 2026 _ Marzo.xlsx")
SHEET_NAME = "Hoja 1"

OU_ALUMNOS = "/Alumnos"
OU_DOCENTE = "/Equipo Docente"

# Partículas que forman parte del apellido compuesto
SURNAME_PARTICLES = {"de", "del", "da", "dos", "las", "los", "la"}

# Columnas (0-indexed)
COL_NOMBRE    = 0  # A
COL_EMAIL     = 1  # B
COL_CURSO     = 2  # C
COL_NUEVO_CORREO = 3  # D (salida)
COL_NUEVA_CONTRA = 4  # E (salida)
COL_MIGRADO      = 5  # F (salida)


# ─── PARSEO DEL NOMBRE ───────────────────────────────────────────────────────

def parse_nombre_apellido(fullname: str) -> tuple:
    """
    Extrae (primer_nombre, primer_apellido) del nombre completo.

    Reglas:
      - 2 palabras: "Iván Funes" → ("Iván", "Funes")
      - 3 palabras: "Brian Barreto Maquia" → ("Brian", "Barreto")
        1 nombre + 2 apellidos → primer nombre + primer apellido
      - 4 palabras: "María José Cáceres Ojeda" → ("María", "Cáceres")
        2 nombres + 2 apellidos → primer nombre + primer apellido (index 2)
      - Partículas: "Mayra de Souza Arbiza" → ("Mayra", "deSouza")
        Si la 2da palabra es partícula ("de"), el apellido = partícula + siguiente
    """
    parts = fullname.strip().split()

    if len(parts) < 2:
        return parts[0], parts[0]

    nombre = parts[0]

    if len(parts) == 2:
        # "Iván Funes"
        return nombre, parts[1]

    if len(parts) == 3:
        # Verificar si parts[1] es partícula: "Mayra de Souza" (caso improbable con 3)
        if parts[1].lower() in SURNAME_PARTICLES:
            return nombre, parts[1].lower() + parts[2]
        # "Brian Barreto Maquia" → 1 nombre + 2 apellidos
        return nombre, parts[1]

    # 4+ palabras
    if parts[1].lower() in SURNAME_PARTICLES:
        # "Mayra de Souza Arbiza" → apellido = "deSouza"
        return nombre, parts[1].lower() + parts[2]

    if parts[2].lower() in SURNAME_PARTICLES:
        # Caso: "María José de Souza Arbiza" → apellido = "deSouza"
        return nombre, parts[2].lower() + parts[3]

    # "María José Cáceres Ojeda" → 2 nombres + 2 apellidos
    return nombre, parts[2]


def generate_email(firstname: str, lastname: str) -> str:
    """Genera email: nombre.apellido@ctcsalto.edu.uy (sin tildes, minúsculas)."""
    first = strip_accents(firstname.strip()).lower()
    last = strip_accents(lastname.strip()).lower()
    return f"{first}.{last}@{DOMAIN}"


# ─── LECTURA DEL XLSX ────────────────────────────────────────────────────────

def read_personas() -> list:
    """
    Lee el xlsx y retorna lista de dicts con estudiantes y docente.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    personas = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or len(row) < 2:
            continue

        nombre = str(row[COL_NOMBRE]).strip() if row[COL_NOMBRE] else ""
        email = str(row[COL_EMAIL]).strip() if row[COL_EMAIL] else ""
        curso = str(row[COL_CURSO]).strip() if len(row) > COL_CURSO and row[COL_CURSO] else ""

        # Skip filas vacías, headers, separadores
        if not nombre or not email:
            continue
        if nombre.lower() in ("estudiantes", "docente", "moodle", "none", "cursos 2026"):
            continue

        # Verificar si ya fue migrado
        migrado = ""
        if len(row) > COL_MIGRADO and row[COL_MIGRADO]:
            migrado = str(row[COL_MIGRADO]).strip().lower()
        if migrado == "si":
            continue

        # Determinar rol por posición (docente está después de fila 20)
        es_docente = i >= 21
        ou = OU_DOCENTE if es_docente else OU_ALUMNOS
        rol = "Docente" if es_docente else "Estudiante"

        firstname, lastname = parse_nombre_apellido(nombre)

        personas.append({
            "row": i,
            "nombre_completo": nombre,
            "firstname": firstname,
            "lastname": lastname,
            "email_personal": email,
            "curso": curso,
            "ou": ou,
            "rol": rol,
        })

    wb.close()
    return personas


# ─── ESCRIBIR RESULTADOS ─────────────────────────────────────────────────────

def write_result(row_number: int, nuevo_correo: str, nueva_contrasena: str):
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    ws.cell(row=row_number, column=COL_NUEVO_CORREO + 1, value=nuevo_correo)
    ws.cell(row=row_number, column=COL_NUEVA_CONTRA + 1, value=nueva_contrasena)
    ws.cell(row=row_number, column=COL_MIGRADO + 1, value="Si")
    wb.save(XLSX_PATH)
    wb.close()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migrar Cursos Marzo 2026 a Google Workspace")
    parser.add_argument("--dry-run", help="Solo muestra el plan, no crea nada", action="store_true")
    args = parser.parse_args()

    print(f"Leyendo {XLSX_PATH}...")
    personas = read_personas()
    print(f"Personas elegibles: {len(personas)}\n")

    if not personas:
        print("No hay personas pendientes de migrar.")
        return

    # ── Plan ──
    print(f"{'=' * 70}")
    print("PLAN DE MIGRACIÓN — CURSOS MARZO 2026")
    print(f"{'=' * 70}")

    seen_emails = set()
    duplicados = []

    for p in personas:
        email_ws = generate_email(p["firstname"], p["lastname"])
        dup = ""
        if email_ws in seen_emails:
            dup = "  ⚠ DUPLICADO"
            duplicados.append((p["nombre_completo"], email_ws))
        seen_emails.add(email_ws)

        print(f"  [{p['row']:>3}] {p['nombre_completo']}  ({p['rol']})")
        print(f"        → {email_ws}  OU={p['ou']}")
        print(f"        Nombre: {p['firstname']} | Apellido: {p['lastname']}")
        print(f"        Email personal: {p['email_personal']}")
        print(f"        Curso: {p['curso']}")
        print()

    if duplicados:
        print(f"\n⚠ DUPLICADOS DETECTADOS ({len(duplicados)}):")
        for nombre, email in duplicados:
            print(f"  - {nombre} → {email}")
        print()

    if args.dry_run:
        print("[DRY-RUN] No se creó nada.")
        return

    # ── Ejecución ──
    print(f"{'=' * 70}")
    print("CREANDO CUENTAS")
    print(f"{'=' * 70}\n")

    service = GoogleWorkspaceService()
    created = 0
    errors = 0

    for p in personas:
        row_num = p["row"]
        firstname = p["firstname"]
        lastname = p["lastname"]
        email_personal = p["email_personal"]
        ou = p["ou"]

        nuevo_correo = generate_email(firstname, lastname)
        nueva_contrasena = generate_password()

        given_name = strip_accents(firstname)
        family_name = strip_accents(lastname)

        print(f"  [{row_num:>3}] {p['nombre_completo']} → {nuevo_correo} (OU={ou})")

        # Etapa 1: crear cuenta
        account_created = False
        try:
            result = service.create_google_account(
                primary_email=nuevo_correo,
                given_name=given_name,
                family_name=family_name,
                password=nueva_contrasena,
                org_unit_path=ou,
            )
            print(f"        [OK] Cuenta creada")
            account_created = True
        except ValueError as e:
            err_msg = str(e)
            if "409" in err_msg or "already exists" in err_msg.lower() or "ya existe" in err_msg.lower():
                print("        [WARN] Cuenta ya existe, continuando...")
                account_created = True
            else:
                print(f"        [ERR] Crear cuenta: {err_msg}")
                errors += 1
                continue
        except Exception as e:
            print(f"        [ERR] Crear cuenta: {e}")
            traceback.print_exc()
            errors += 1
            continue

        if not account_created:
            errors += 1
            continue

        # Etapa 2: notificación
        try:
            send_notification(
                nuevo_correo=nuevo_correo,
                nueva_contrasena=nueva_contrasena,
                firstname=firstname,
                lastname=lastname,
                email_original=email_personal,
            )
            print("        [OK] Notificación enviada")
        except Exception as e:
            print(f"        [ERR] Notificación: {e}")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            errors += 1
            continue

        # Etapa 3: guardar en xlsx
        try:
            write_result(row_num, nuevo_correo, nueva_contrasena)
            print(f"        [OK] xlsx actualizado (fila {row_num})")
            created += 1
        except Exception as e:
            print(f"        [ERR] Escribir xlsx: {e}")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            errors += 1

    # ── Reporte ──
    print(f"\n{'=' * 70}")
    print("REPORTE — CURSOS MARZO 2026")
    print(f"{'=' * 70}")
    print(f"  Cuentas creadas: {created}")
    print(f"  Errores:         {errors}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
