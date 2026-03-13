"""
Script de migración de NUEVOS INSCRIPTOS desde Excel a Google Workspace.

Lee un xlsx con estructura simple (columna A = nombre completo, columna B = email personal).
Para cada alumno:
  1. Parsea nombre y apellido desde el nombre completo
  2. Genera email: primer_nombre.apellido@ctcsalto.edu.uy
  3. Crea cuenta en Google Workspace (OU=/Alumnos)
  4. Envía notificación por n8n al email personal
  5. Escribe Nuevo Correo, Nueva Contraseña y Migrado=Si en el xlsx

Parseo del nombre completo:
  - 2 palabras: "Rafel Lupano" → nombre=Rafel, apellido=Lupano
  - 3+ palabras: "Bruno Nahuel Barboza" → nombre=Bruno, apellido=Barboza
    (usa primer nombre + último token como apellido)

Manejo de errores: idéntico a migrate_usuarios.py
  - Si la cuenta ya existe (409) → continúa con notificación
  - Si la notificación falla → NO marca migrado, al re-ejecutar reintenta
  - Si falla xlsx → imprime credenciales en consola

Uso:
    python migrate_nuevos_inscriptos.py --dry-run          # ver plan sin ejecutar
    python migrate_nuevos_inscriptos.py --alumno "Rafel Lupano"  # solo ese alumno
    python migrate_nuevos_inscriptos.py                    # todos los pendientes
"""

import sys
import os
import argparse
import traceback

# ── sys.path → raíz del proyecto (3 niveles arriba de scripts/) ──────────────
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, PROJECT_ROOT)

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"), override=True)

import openpyxl
from external_services.google.google_service import GoogleWorkspaceService

# Reutilizar utilidades del script de migración existente
from migrate_usuarios import (
    strip_accents,
    _extract_surname,
    generate_password,
    send_notification,
    DOMAIN,
)

# ─── CONFIG ───────────────────────────────────────────────────────────────────
XLSX_PATH = os.path.join(PROJECT_ROOT, "Nuevos Inscriptos.xlsx")
SHEET_NAME = "Hoja1"
OU_ALUMNOS = "/Alumnos"

# Filas de datos empiezan después de los headers
# Fila 5 = headers ("Estudiantes", "Mail"), datos desde fila 6
HEADER_ROW = 5

# Columnas (0-indexed)
COL_NOMBRE_COMPLETO = 0   # A
COL_EMAIL_PERSONAL  = 1   # B
# Columnas de salida (se crean/escriben por el script)
COL_NUEVO_CORREO    = 2   # C
COL_NUEVA_CONTRA    = 3   # D
COL_MIGRADO         = 4   # E


# ─── PARSEO DEL NOMBRE COMPLETO ──────────────────────────────────────────────

def parse_fullname(fullname: str) -> tuple:
    """
    Parsea nombre completo en (primer_nombre, apellido).

    Reglas:
      - 2 palabras: "Rafel Lupano" → ("Rafel", "Lupano")
      - 3+ palabras: "Bruno Nahuel Barboza" → ("Bruno", "Barboza")
        Toma primer token como nombre, último token como apellido.
      - 1 palabra: retorna (palabra, palabra) como fallback

    Returns:
        (primer_nombre, apellido)
    """
    parts = fullname.strip().split()
    if len(parts) >= 3:
        return parts[0], parts[-1]
    elif len(parts) == 2:
        return parts[0], parts[1]
    else:
        # Fallback: 1 sola palabra
        return parts[0], parts[0]


def generate_workspace_email(firstname: str, lastname: str) -> str:
    """Genera email: primer_nombre.apellido@ctcsalto.edu.uy (sin tildes, minúsculas)."""
    first = strip_accents(firstname.strip()).lower()
    last  = strip_accents(lastname.strip()).lower()
    return f"{first}.{last}@{DOMAIN}"


# ─── LECTURA DEL XLSX ────────────────────────────────────────────────────────

def read_nuevos_inscriptos(alumno_filter: str = None) -> list:
    """
    Lee la Hoja1 del xlsx y retorna lista de dicts con los alumnos elegibles.
    Skipea: filas vacías, headers, filas sin email, ya migrados.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]

    alumnos = []
    for i, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        nombre_completo = str(row[COL_NOMBRE_COMPLETO]).strip() if row[COL_NOMBRE_COMPLETO] else ""
        email_personal  = str(row[COL_EMAIL_PERSONAL]).strip()  if row[COL_EMAIL_PERSONAL]  else ""

        # Skip filas vacías o sin datos útiles
        if not nombre_completo or not email_personal:
            continue

        # Skip si parece header repetido
        if nombre_completo.lower() in ("estudiantes", "moodle", "none"):
            continue

        # Verificar si ya fue migrado (columna E)
        migrado = ""
        if len(row) > COL_MIGRADO and row[COL_MIGRADO]:
            migrado = str(row[COL_MIGRADO]).strip().lower()

        if migrado == "si":
            continue

        # Parsear nombre
        firstname, lastname = parse_fullname(nombre_completo)

        # Filtro por alumno específico
        if alumno_filter:
            if alumno_filter.lower() not in nombre_completo.lower():
                continue

        alumnos.append({
            "row":              i,
            "nombre_completo":  nombre_completo,
            "firstname":        firstname,
            "lastname":         lastname,
            "email_personal":   email_personal,
        })

    wb.close()
    return alumnos


# ─── ESCRIBIR RESULTADOS AL XLSX ─────────────────────────────────────────────

def write_migration_result(row_number: int, nuevo_correo: str, nueva_contrasena: str):
    """
    Escribe Nuevo Correo (C), Nueva Contraseña (D) y Migrado=Si (E) en la fila.
    """
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    ws.cell(row=row_number, column=COL_NUEVO_CORREO + 1, value=nuevo_correo)
    ws.cell(row=row_number, column=COL_NUEVA_CONTRA + 1, value=nueva_contrasena)
    ws.cell(row=row_number, column=COL_MIGRADO + 1,      value="Si")

    wb.save(XLSX_PATH)
    wb.close()


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Migrar nuevos inscriptos a Google Workspace")
    parser.add_argument("--alumno",  help="Filtrar por nombre (parcial)", default=None)
    parser.add_argument("--dry-run", help="Solo muestra el plan, no crea nada", action="store_true")
    args = parser.parse_args()

    # ── Lectura ──
    print(f"Leyendo {XLSX_PATH}...")
    alumnos = read_nuevos_inscriptos(alumno_filter=args.alumno)
    print(f"Alumnos elegibles: {len(alumnos)}\n")

    if not alumnos:
        print("No hay alumnos pendientes de migrar.")
        return

    # ── Plan ──
    print(f"{'=' * 70}")
    print("PLAN DE MIGRACIÓN — NUEVOS INSCRIPTOS")
    print(f"{'=' * 70}")

    seen_emails = set()
    duplicados = []

    for a in alumnos:
        email_ws = generate_workspace_email(a["firstname"], a["lastname"])
        dup = ""
        if email_ws in seen_emails:
            dup = "  ⚠ DUPLICADO"
            duplicados.append((a["nombre_completo"], email_ws))
        seen_emails.add(email_ws)

        print(f"  [{a['row']:>3}] {a['nombre_completo']}")
        print(f"        → {email_ws}  OU={OU_ALUMNOS}{dup}")
        print(f"        Nombre: {a['firstname']} | Apellido: {a['lastname']}")
        print(f"        Email personal: {a['email_personal']}")
        print()

    if duplicados:
        print(f"{'=' * 70}")
        print(f"⚠ DUPLICADOS DETECTADOS ({len(duplicados)}):")
        for nombre, email in duplicados:
            print(f"  - {nombre} → {email}")
        print(f"{'=' * 70}")
        print()

    if args.dry_run:
        print("[DRY-RUN] No se creó nada.")
        return

    # ── Ejecución ──
    print(f"{'=' * 70}")
    print("CREANDO CUENTAS")
    print(f"{'=' * 70}\n")

    service = GoogleWorkspaceService()
    created = 0
    errors  = 0

    for a in alumnos:
        row_num        = a["row"]
        firstname      = a["firstname"]
        lastname       = a["lastname"]
        email_personal = a["email_personal"]
        nombre_display = a["nombre_completo"]

        nuevo_correo     = generate_workspace_email(firstname, lastname)
        nueva_contrasena = generate_password()

        # given_name y family_name para Google Workspace (sin tildes para evitar rechazo)
        given_name  = strip_accents(firstname)
        family_name = strip_accents(lastname)

        print(f"  [{row_num:>3}] {nombre_display} → {nuevo_correo}")

        # ─── ETAPA 1: crear cuenta Google Workspace ──────────────────────
        account_created = False
        try:
            result = service.create_google_account(
                primary_email=nuevo_correo,
                given_name=given_name,
                family_name=family_name,
                password=nueva_contrasena,
                org_unit_path=OU_ALUMNOS,
            )
            print(f"        [OK] Cuenta creada: {result}")
            account_created = True

        except ValueError as e:
            err_msg = str(e)
            if "409" in err_msg or "already exists" in err_msg.lower() or "ya existe" in err_msg.lower():
                print("        [WARN] Cuenta ya existe en Google Workspace, continuando...")
                account_created = True
            else:
                print(f"        [ERR] Etapa 1 - Crear cuenta: {err_msg}")
                errors += 1
                continue

        except Exception as e:
            print(f"        [ERR] Etapa 1 - Crear cuenta: {e}")
            traceback.print_exc()
            errors += 1
            continue

        # ─── ETAPA 2: enviar notificación ────────────────────────────────
        if not account_created:
            print("        [ERR] Cuenta no creada, skipeando notificación.")
            errors += 1
            continue

        try:
            send_notification(
                nuevo_correo=nuevo_correo,
                nueva_contrasena=nueva_contrasena,
                firstname=firstname,
                lastname=lastname,
                email_original=email_personal,
            )
            print("        [OK] Notificación enviada")

        except Exception as e:
            print(f"        [ERR] Etapa 2 - Notificación: {e}")
            print("        ⚠ CUENTA YA CREADA pero notificación falló.")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            print("        ⚠ Re-ejecutar el script para reintentar la notificación.")
            errors += 1
            continue

        # ─── ETAPA 3: guardar en xlsx ────────────────────────────────────
        try:
            write_migration_result(row_num, nuevo_correo, nueva_contrasena)
            print(f"        [OK] xlsx actualizado (fila {row_num}, Migrado=Si)")
            created += 1

        except Exception as e:
            print(f"        [ERR] Etapa 3 - Escribir xlsx: {e}")
            print("        ⚠ CUENTA CREADA Y NOTIFICADA, pero xlsx no se actualizó.")
            print(f"        ⚠ Credenciales: {nuevo_correo} / {nueva_contrasena}")
            errors += 1

    # ── Reporte final ──
    print(f"\n{'=' * 70}")
    print("REPORTE DE MIGRACIÓN — NUEVOS INSCRIPTOS")
    print(f"{'=' * 70}")
    print(f"  Cuentas creadas exitosas: {created}")
    print(f"  Errores:                  {errors}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
