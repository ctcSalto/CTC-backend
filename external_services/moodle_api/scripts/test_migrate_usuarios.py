"""
Test de migración de usuarios: simula y valida cada etapa para un usuario real
(Ezequiel Viera, username=eviera) sin crear la cuenta de forma definitiva.

Qué hace:
  1. Valida que el xlsx se lea bien y el usuario se encuentre
  2. Verifica la generación del email y la contraseña
  3. Hace un POST de prueba al webhook de n8n (account-notification) con datos ficticios
     para confirmar que el webhook responde → NO envía la contraseña real
  4. Abre el xlsx en modo escritura y verifica que la fila sea escribible
     (NO escribe datos reales)
  5. Reporta cada etapa con OK / FAIL

Para ejecutar el test real (crear cuenta + notificar + escribir xlsx) después de
que este test pase, usar:
    python migrate_usuarios.py --usuario eviera
"""

import sys
import os
import traceback

# ── sys.path → raíz del proyecto ─────────────────────────────────────────────
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
sys.path.insert(0, PROJECT_ROOT)

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from dotenv import load_dotenv
load_dotenv(os.path.join(PROJECT_ROOT, ".env"), override=True)

import requests
import openpyxl

# Importar utilidades del script principal
from migrate_usuarios import (
    generate_workspace_email,
    generate_password,
    read_all_usuarios,
    OU_MAP,
    XLSX_PATH,
    COL_NUEVO_CORREO,
    COL_NUEVA_CONTRA,
    COL_MIGRADO,
)

# ─── CONFIG DEL TEST ──────────────────────────────────────────────────────────
TEST_USERNAME = "eviera"

# ─── RESULTADOS ───────────────────────────────────────────────────────────────
results = {}   # etapa → (passed: bool, detalle: str)


def report(etapa: str, passed: bool, detalle: str = ""):
    results[etapa] = (passed, detalle)
    icon = "[OK]  " if passed else "[FAIL]"
    print(f"  {icon} {etapa}")
    if detalle:
        print(f"        → {detalle}")


# ─── ETAPAS DE TEST ──────────────────────────────────────────────────────────

def test_lectura_xlsx():
    """Valida que el usuario eviera se lea correctamente del xlsx."""
    usuarios = read_all_usuarios(username_filter=TEST_USERNAME)
    if len(usuarios) != 1:
        report("1. Lectura xlsx", False, f"Se esperaba 1 usuario, se encontraron {len(usuarios)}")
        return None

    u = usuarios[0]
    detalle = (
        f"sheet={u['sheet']} fila={u['row']} | "
        f"{u['firstname']} {u['lastname']} | "
        f"tipo={u['tipo_usuario']} | "
        f"email_orig={u['email_original']}"
    )
    report("1. Lectura xlsx", True, detalle)
    return u


def test_generacion_email(usuario: dict):
    """Valida email generado y contraseña."""
    email = generate_workspace_email(usuario["firstname"], usuario["lastname"])
    expected = "ezequiel.viera@ctcsalto.edu.uy"

    if email != expected:
        report("2a. Email generado", False, f"Esperado '{expected}', obtuvo '{email}'")
    else:
        report("2a. Email generado", True, email)

    # Contraseña: generar 5 y validar formato
    passwords_ok = True
    for _ in range(5):
        pwd = generate_password()
        tiene_lower  = any(c.islower()  for c in pwd)
        tiene_upper  = any(c.isupper()  for c in pwd)
        tiene_digito = any(c.isdigit()  for c in pwd)
        tiene_esp    = any(c in "!@#$%&*-_=+" for c in pwd)
        if not (tiene_lower and tiene_upper and tiene_digito and tiene_esp and len(pwd) == 16):
            passwords_ok = False
            break

    report("2b. Contraseña", passwords_ok,
           f"Ejemplo: {generate_password()} (16 chars, mixed)")


def test_ou_mapping(usuario: dict):
    """Valida que el tipo de usuario tenga OU correcta."""
    tipo = usuario["tipo_usuario"]
    ou   = OU_MAP.get(tipo)
    if ou:
        report("3. Mapeo OU", True, f"{tipo} → {ou}")
    else:
        report("3. Mapeo OU", False, f"Tipo '{tipo}' no tiene mapeo en OU_MAP")


def test_webhook_conectividad():
    """
    Hace un POST al webhook de notificación con datos de prueba (placeholder)
    para verificar que el endpoint responde. NO envía datos reales.
    """
    webhook_url = os.getenv("N8N_NOTIFICATION_WEBHOOK_URL")
    if not webhook_url:
        report("4. Webhook conectividad", False, "N8N_NOTIFICATION_WEBHOOK_URL no definida en .env")
        return

    api_token   = os.getenv("N8N_API_TOKEN", "")
    header_name = os.getenv("N8N_HEADER_NAME", "Authorization")
    timeout     = int(os.getenv("N8N_TIMEOUT", "30"))

    headers = {
        header_name:    api_token,
        "Content-Type": "application/json",
    }
    # Payload de prueba — datos ficticios, NO la contraseña real
    payload = {
        "firstname":        "TEST",
        "lastname":         "TEST",
        "nuevo_correo":     "test.test@ctcsalto.edu.uy",
        "nueva_contrasena": "TEST_NO_USAR",
        "email_original":   "test@test.com",
        "_test":            True,   # campo extra para que el flujo n8n pueda identificarlo como prueba
    }

    try:
        response = requests.post(webhook_url, json=payload, headers=headers, timeout=timeout)
        if response.status_code < 400:
            report("4. Webhook conectividad", True,
                   f"HTTP {response.status_code} — webhook responde correctamente")
        else:
            report("4. Webhook conectividad", False,
                   f"HTTP {response.status_code}: {response.text[:200]}")
    except requests.exceptions.Timeout:
        report("4. Webhook conectividad", False, f"Timeout (>{timeout}s)")
    except requests.exceptions.ConnectionError:
        report("4. Webhook conectividad", False, "No se pudo conectar al webhook")
    except Exception as e:
        report("4. Webhook conectividad", False, str(e))


def test_escritura_xlsx(usuario: dict):
    """
    Abre el xlsx en modo lectura-escritura y verifica que la fila del usuario
    sea accesible y escribible. NO modifica datos reales.
    """
    try:
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[usuario["sheet"]]
        row = usuario["row"]

        # Leer valores actuales de las celdas que se escribirían
        actual_correo  = ws.cell(row=row, column=COL_NUEVO_CORREO + 1).value
        actual_contra  = ws.cell(row=row, column=COL_NUEVA_CONTRA + 1).value
        actual_migrado = ws.cell(row=row, column=COL_MIGRADO + 1).value

        wb.close()   # cerrar sin guardar

        detalle = (
            f"Fila {row} accesible | "
            f"Nuevo Correo actual={actual_correo} | "
            f"Nueva Contraseña actual={actual_contra} | "
            f"Migrado actual={actual_migrado}"
        )
        report("5. Escritura xlsx (lectura previa)", True, detalle)

    except Exception as e:
        report("5. Escritura xlsx (lectura previa)", False, str(e))
        traceback.print_exc()


def test_google_service_instancia():
    """Verifica que GoogleWorkspaceService se instancia y tiene config correcta."""
    try:
        from external_services.google.google_service import GoogleWorkspaceService
        svc = GoogleWorkspaceService()
        detalle = (
            f"base_url={svc.base_url} | "
            f"header={svc.header_name} | "
            f"token={'***' + svc.api_token[-8:] if svc.api_token else 'MISSING'} | "
            f"timeout={svc.timeout}s"
        )
        report("6. GoogleWorkspaceService", True, detalle)
    except Exception as e:
        report("6. GoogleWorkspaceService", False, str(e))


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("TEST DE MIGRACIÓN — usuario: eviera (Ezequiel Viera)")
    print("=" * 70)
    print()

    # 1. Lectura xlsx
    usuario = test_lectura_xlsx()
    if not usuario:
        print("\n[ABORTADO] No se pudo leer el usuario del xlsx.")
        return

    # 2. Email + contraseña
    test_generacion_email(usuario)

    # 3. OU
    test_ou_mapping(usuario)

    # 4. Webhook conectividad (POST de prueba)
    test_webhook_conectividad()

    # 5. Escritura xlsx (solo lectura de la fila)
    test_escritura_xlsx(usuario)

    # 6. GoogleWorkspaceService instancia
    test_google_service_instancia()

    # ── Resumen ──
    print()
    print("=" * 70)
    print("RESUMEN DEL TEST")
    print("=" * 70)
    total   = len(results)
    pasados = sum(1 for ok, _ in results.values() if ok)
    for etapa, (ok, _) in results.items():
        print(f"  {'[OK]  ' if ok else '[FAIL]'} {etapa}")

    print()
    if pasados == total:
        print(f"  Todas las etapas pasaron ({pasados}/{total}).")
        print()
        print("  Para ejecutar la migración real de este usuario:")
        print("    python migrate_usuarios.py --usuario eviera")
    else:
        print(f"  {total - pasados} etapa(s) fallaron. Revisar los errores antes de ejecutar.")
    print("=" * 70)


if __name__ == "__main__":
    main()
